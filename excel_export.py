"""
excel_export.py — write the formatted workbook: Papers, Analysis (with three
charts), FY Trend, Keywords, Search Log, Dedup Review, Not Relevant (Backup).
"""

import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, LineChart, PieChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.legend import Legend
from openpyxl.chart.layout import Layout, ManualLayout
from openpyxl.chart.marker import DataPoint

from fiscal_year import current_and_prior_fy, fy_to_year
from matching import build_search_query
from normalize import clean_cell
from keywords import SURVEY_FAMILIES, iter_terms
from relevance import IDENTITY_MIN, USE_MIN

OUTPUT_COLS = [
    "title", "doi", "year", "month", "pub_date", "fy",
    "publication_type", "pub_type", "journal_tier", "peer_reviewed_auto", "venue",
    "authors", "first_author", "second_author", "n_authors",
    "link", "open_access", "oa_url",
    "wb_affiliation_auto", "multilateral_affiliation", "university_affiliation",
    "affiliations", "affiliation_countries",
    "geography_clean", "is_first_author_africa", "is_any_author_africa",
    "is_africa_institution_strict",
    "identity_score", "use_score", "relevance_score", "relevance_flags",
    "match_tier", "match_reason",
    "dataset_country", "research_topics", "abstract",
    "citation_count", "language",
    "survey_family", "survey_terms_matched", "openalex_id", "source",
    "date_discovered",
]

COL_WIDTHS = {
    "title": 60, "doi": 34, "abstract": 60,
    "authors": 35, "affiliations": 40, "affiliation_countries": 20,
    "university_affiliation": 35, "multilateral_affiliation": 25,
    "venue": 25, "survey_family": 30, "survey_terms_matched": 38,
    "relevance_flags": 30, "match_reason": 34, "geography_clean": 20,
}

C_DARK  = "1F5C99"
C_HL    = "C5D9F1"
C_WHITE = "FFFFFF"

_THIN_BLACK = Side(style="thin", color="000000")
_BORDER_ALL = Border(left=_THIN_BLACK, right=_THIN_BLACK,
                     top=_THIN_BLACK, bottom=_THIN_BLACK)
_F_WHITE_BOLD = Font(bold=True, color=C_WHITE)
_F_BOLD       = Font(bold=True)
_F_BOLD_BLACK = Font(bold=True, color="000000")
_FILL_DARK    = PatternFill("solid", start_color=C_DARK, end_color=C_DARK)
_FILL_HL      = PatternFill("solid", start_color=C_HL, end_color=C_HL)
_AL_CENTER    = Alignment(horizontal="center")


def _build_trend_df(papers: list, current_fy: str) -> pd.DataFrame:
    fy_labels = [f"FY{str(y)[-2:]}" for y in range(2009, 2030)]
    rows = []
    for fy in fy_labels:
        fp = [p for p in papers if p.get("fy") == fy]
        rows.append((fy, len(fp),
                     sum(1 for p in fp if p.get("is_any_author_africa")),
                     sum(1 for p in fp if p.get("is_first_author_africa")),
                     sum(1 for p in fp if (p.get("peer_reviewed_auto") or "") == "Yes"),
                     sum(1 for p in fp if (p.get("wb_affiliation_auto") or "") == "Yes")))
    return pd.DataFrame(rows, columns=["FY", "All Papers", "Africa (any author)",
                                       "Africa (1st author)", "Peer-Reviewed", "WB-Affiliated"])


def _cell(ws, r, c, value=None, bold=False, white_text=False, fill=None, pct=False):
    cell = ws.cell(r, c)
    if value is not None:
        cell.value = value
    if bold or white_text:
        cell.font = Font(bold=bold, color=C_WHITE if white_text else "000000")
    if fill:
        cell.fill = fill
    if pct and isinstance(value, (int, float)):
        cell.number_format = "0%"
    return cell


def _section_row(ws, r, label):
    c = ws.cell(r, 1, label)
    c.font = _F_WHITE_BOLD
    c.fill = _FILL_DARK
    c.border = _BORDER_ALL
    for col in range(2, 5):
        cc = ws.cell(r, col)
        cc.fill = _FILL_DARK
        cc.border = _BORDER_ALL
    ws.merge_cells(f"A{r}:D{r}")


def _highlight(cell, pct=False):
    cell.fill = _FILL_HL
    cell.font = _F_BOLD_BLACK
    cell.border = _BORDER_ALL
    if pct:
        cell.number_format = "0%"


def _data_row(ws, r, label, note="", count=None, share=None,
              hl_count=False, hl_share=False, bold_label=False):
    lc = ws.cell(r, 1, label)
    if bold_label:
        lc.font = _F_BOLD
    ws.cell(r, 2, note)
    if count is not None:
        cc = ws.cell(r, 3, count)
        if hl_count:
            _highlight(cc)
        elif bold_label:
            cc.font = _F_BOLD
    if share is not None:
        sc = ws.cell(r, 4, share)
        sc.number_format = "0%"
        if hl_share:
            _highlight(sc, pct=True)


def write_analysis_sheet(ws, papers, current_fy, completed_fys,
                         min_year, excluded, trend_df):
    excluded = excluded or []
    total  = len(papers)
    # why each excluded paper was set aside -- the two axes fail for very
    # different reasons and lumping them together hides which is doing the work
    n_no_use   = sum(1 for p in excluded
                    if (p.get("use_score") or 0) < USE_MIN
                    and (p.get("identity_score") or 0) >= IDENTITY_MIN)
    n_no_ident = sum(1 for p in excluded if (p.get("identity_score") or 0) < IDENTITY_MIN)
    n_vetoed   = sum(1 for p in excluded
                    if "excluded_pub_type" in (p.get("relevance_flags") or ""))
    peer   = sum(1 for p in papers if (p.get("peer_reviewed_auto") or "") == "Yes")
    wb_    = sum(1 for p in papers if (p.get("wb_affiliation_auto") or "") == "Yes")
    mult   = sum(1 for p in papers if p.get("multilateral_affiliation"))
    af_any = sum(1 for p in papers if p.get("is_any_author_africa"))
    af_1st = sum(1 for p in papers if p.get("is_first_author_africa"))
    af_str = sum(1 for p in papers if p.get("is_africa_institution_strict"))
    unk    = sum(1 for p in papers if p.get("geography_clean") == "Unclassified")
    r_border = sum(1 for p in papers
                  if (p.get("identity_score") or 0) == IDENTITY_MIN
                  and (p.get("use_score") or 0) == USE_MIN)
    r_backed = total - r_border
    r_use    = sum(1 for p in papers if (p.get("use_score") or 0) > USE_MIN)
    # match_tier carries EVERY tier a paper matched on ("A; C" = it matched an
    # unambiguous name and an acronym). Counting membership per tier therefore
    # double-counts and sums well past 100%, which is just confusing. Report
    # the strongest tier each paper reached -- mutually exclusive, sums to
    # 100% -- and give the overlap its own line.
    def _best_tier(p) -> str:
        t = p.get("match_tier") or ""
        return "A" if "A" in t else "B" if "B" in t else "C" if "C" in t else ""
    tA = sum(1 for p in papers if _best_tier(p) == "A")
    tB = sum(1 for p in papers if _best_tier(p) == "B")
    tC = sum(1 for p in papers if _best_tier(p) == "C")
    t_multi = sum(1 for p in papers
                 if len([x for x in (p.get("match_tier") or "").split(";") if x.strip()]) > 1)

    fy0        = completed_fys[0]
    fy0_total  = sum(1 for p in papers if p.get("fy") == fy0)
    af_any_fy0 = sum(1 for p in papers if p.get("fy") == fy0 and p.get("is_any_author_africa"))
    af_1st_fy0 = sum(1 for p in papers if p.get("fy") == fy0 and p.get("is_first_author_africa"))

    pct     = lambda n: (n / total) if total else 0
    pct_fy0 = lambda n: (n / fy0_total) if fy0_total else 0

    for c, h in enumerate(["Metric", "Note", "Count", "Share"], 1):
        _cell(ws, 1, c, h, bold=True, white_text=True, fill=_FILL_DARK)
    ws.cell(1, 3).alignment = _AL_CENTER
    ws.cell(1, 4).alignment = _AL_CENTER
    r = 2

    _section_row(ws, r, "── TOTALS (papers that both identify AND use the survey) ──"); r += 1
    _data_row(ws, r, "TOTAL PAPERS ANALYSED", "all figures below refer to this set",
              total, bold_label=True); r += 1
    _data_row(ws, r, "  Excluded — mentions the survey but no evidence it used the data",
              "in 'Not Relevant (Backup)' sheet", n_no_use); r += 1
    _data_row(ws, r, "  Excluded — not confidently our survey (name collision / other country)",
              "in 'Not Relevant (Backup)' sheet", n_no_ident); r += 1
    _data_row(ws, r, "  Excluded — publication type can't be an empirical use",
              "conference abstract, dataset, paratext, erratum, letter, software", n_vetoed); r += 1
    _data_row(ws, r, "  Total retrieved before exclusion", "", total + len(excluded)); r += 1
    _data_row(ws, r, "Peer-reviewed journal articles (auto-detected)", "", peer, pct(peer)); r += 1
    _data_row(ws, r, "World Bank–affiliated papers (auto-detected)", "", wb_, pct(wb_)); r += 1
    _data_row(ws, r, "Multilateral org–affiliated (IFPRI/FAO/CGIAR…)", "", mult, pct(mult)); r += 1
    r += 1

    _section_row(ws, r, "── FLOW: new papers by WBG fiscal year ──────────────"); r += 1
    flow_cur = sum(1 for p in papers if p.get("fy") == current_fy)
    _data_row(ws, r, f"  {current_fy}", "current FY — still in progress", flow_cur); r += 1
    for i, fy in enumerate(completed_fys):
        n = sum(1 for p in papers if p.get("fy") == fy)
        note = "◄ most recently completed FY" if i == 0 else ""
        _data_row(ws, r, f"  {fy}", note, n, hl_count=(i == 0)); r += 1
    r += 1

    _section_row(ws, r, "── GATE 1: HOW EACH PAPER WAS MATCHED ─────────────"); r += 1
    _data_row(ws, r, "  Strongest tier each paper matched on. A paper often matches several "
                     "terms at different tiers (e.g. 'LSMS-ISA' AND 'LSMS'); it is counted "
                     "once here, under the strongest. These three sum to 100%."); r += 1
    _data_row(ws, r, "  Tier A — unambiguous name (e.g. 'Uganda National Panel Survey')",
              "full-text match accepted on its own", tA, pct(tA)); r += 1
    _data_row(ws, r, "  Tier B — generic survey name (e.g. 'National Panel Survey')",
              "needs a country context word + allowed field", tB, pct(tB)); r += 1
    _data_row(ws, r, "  Tier C — short acronym (e.g. 'IHPS', 'LSMS')",
              "country + field, plus a case check against collisions", tC, pct(tC)); r += 1
    _data_row(ws, r, "  (of which: matched terms at more than one tier)",
              "counted above under the strongest only", t_multi, pct(t_multi)); r += 1
    r += 1

    _section_row(ws, r, "── GATE 2: IDENTITY AND USE, SCORED SEPARATELY ────"); r += 1
    _data_row(ws, r, "  Two questions, deliberately kept apart. IDENTITY: is this really one of "
                     "our surveys, or a name collision? USE: did the authors actually work with "
                     "the microdata, or do they only mention it? A strong signal is worth 2 "
                     f"points, a weak one 1. Identity needs {IDENTITY_MIN}+ AND use needs "
                     f"{USE_MIN}+ — clearing one axis alone is not enough, which is what stops a "
                     "paper that merely name-drops the survey from being counted.", ""); r += 1
    _data_row(ws, r, "  Use evidence must be TIED to the survey", "'we use panel data' proves "
                     "nothing about WHICH panel — only proximity in the full text, or an "
                     "abstract that names the survey AND describes using it, counts fully"); r += 1
    _data_row(ws, r, "  Borderline — both axes exactly at the minimum",
              "just cleared, worth a skim", r_border, pct(r_border)); r += 1
    _data_row(ws, r, "  Well-backed — corroborated beyond the minimum",
              "", r_backed, pct(r_backed)); r += 1
    _data_row(ws, r, f"  Strong data-use evidence (use score > {USE_MIN})",
              "first-person methods language, or cites the microdata catalogue",
              r_use, pct(r_use)); r += 1
    _data_row(ws, r, "  (Everything that failed either axis is in 'Not Relevant (Backup)')",
              "not counted anywhere above"); r += 1
    r += 1

    _section_row(ws, r, "── SHARE: Geography of Authors ──────────────────────────"); r += 1
    _data_row(ws, r, "NOTE: institution-based (undercounts African diaspora at WB/US/EU)"); r += 1
    _data_row(ws, r, "  Any author at African institution", "≈ original SSA + Mixed",
              af_any, pct(af_any), hl_share=True); r += 1
    _data_row(ws, r, f"    of which in {fy0} (share of {fy0} papers)",
              f"{af_any_fy0} of {fy0_total} {fy0} papers",
              af_any_fy0, pct_fy0(af_any_fy0), hl_share=True); r += 1
    _data_row(ws, r, "  First author at African institution", "", af_1st, pct(af_1st)); r += 1
    _data_row(ws, r, f"    of which in {fy0} (share of {fy0} papers)",
              f"{af_1st_fy0} of {fy0_total} {fy0} papers",
              af_1st_fy0, pct_fy0(af_1st_fy0), hl_share=True); r += 1
    _data_row(ws, r, "  ALL authors at SSA institution", "", af_str, pct(af_str)); r += 1
    _data_row(ws, r, "  Geography unclassified (no OpenAlex institution data)", "",
              unk, pct(unk)); r += 1

    for row_i in range(2, r + 1):
        cell = ws.cell(row_i, 4)
        if isinstance(cell.value, float):
            cell.number_format = "0%"

    ws.column_dimensions["A"].width = 75
    ws.column_dimensions["B"].width = 48
    ws.column_dimensions["C"].width = 23
    ws.column_dimensions["D"].width = 23

    tbl_start = r + 2
    for j, h in enumerate(["Fiscal Year", "Papers",
                           "Africa share (any author)", "Africa share (1st author)"], 1):
        c = ws.cell(tbl_start, j, h)
        c.font = _F_WHITE_BOLD
        c.fill = _FILL_DARK
        c.alignment = _AL_CENTER

    chart_rows = []
    for _, row_d in trend_df.iterrows():
        fy    = str(row_d["FY"])
        n_all = int(row_d["All Papers"])
        n_aa  = int(row_d["Africa (any author)"])
        n_fa  = int(row_d["Africa (1st author)"])
        if n_all == 0 or fy == current_fy:
            continue
        r_i = tbl_start + 1 + len(chart_rows)
        ws.cell(r_i, 1, fy)
        ws.cell(r_i, 2, n_all)
        ws.cell(r_i, 3, n_aa / n_all).number_format = "0%"
        ws.cell(r_i, 4, n_fa / n_all).number_format = "0%"
        chart_rows.append((fy, n_all, n_aa, n_fa))
    tbl_end = tbl_start + len(chart_rows)

    if not chart_rows:
        return

    cats  = Reference(ws, min_col=1, min_row=tbl_start + 1, max_row=tbl_end)
    tot_r = Reference(ws, min_col=2, max_col=2, min_row=tbl_start, max_row=tbl_end)

    bar = BarChart()
    bar.type = "col"
    bar.grouping = "clustered"
    bar.title = "LSMS Papers per Fiscal Year (number of papers)"
    bar.y_axis.title = "Number of papers"
    bar.x_axis.title = "Fiscal Year (World Bank: 1 Jul – 30 Jun)"
    bar.x_axis.tickLblPos = "low"
    bar.y_axis.crosses = "autoZero"
    bar.width, bar.height = 15, 7.5
    bar.legend = None
    bar.add_data(tot_r, titles_from_data=True)
    bar.set_categories(cats)
    bar.series[0].graphicalProperties.solidFill = C_DARK
    bar.series[0].graphicalProperties.line.solidFill = C_DARK
    bar.dataLabels = DataLabelList()
    bar.dataLabels.showVal = True
    bar.dataLabels.showLegendKey = False
    bar.dataLabels.showCatName = False
    bar.dataLabels.showSerName = False
    ws.add_chart(bar, "E1")

    any_r = Reference(ws, min_col=3, max_col=3, min_row=tbl_start, max_row=tbl_end)
    fst_r = Reference(ws, min_col=4, max_col=4, min_row=tbl_start, max_row=tbl_end)

    line = LineChart()
    line.title = "Share of African Scholars per Fiscal Year"
    line.y_axis.title = "Share of papers (%)"
    line.x_axis.title = "Fiscal Year (World Bank: 1 Jul – 30 Jun)"
    line.x_axis.tickLblPos = "low"
    line.y_axis.numFmt = "0%"
    line.y_axis.crosses = "autoZero"
    line.width, line.height = 15, 7.5
    line.style = 10
    line.add_data(any_r, titles_from_data=True)
    line.add_data(fst_r, titles_from_data=True)
    line.set_categories(cats)
    for idx, (clr, sym) in enumerate([("2E7D32", "circle"), ("E65100", "diamond")]):
        s = line.series[idx]
        s.graphicalProperties.line.solidFill = clr
        s.graphicalProperties.line.width = 28000
        s.marker.symbol, s.marker.size = sym, 6
        s.marker.graphicalProperties.solidFill = clr
        s.marker.graphicalProperties.line.solidFill = clr
    # Each point shows only its percentage. Turning on series or category name
    # here is what overlaps the labels into an unreadable block.
    line.dataLabels = DataLabelList()
    line.dataLabels.showVal = True
    line.dataLabels.numFmt = "0%"
    line.dataLabels.showLegendKey = False
    line.dataLabels.showCatName = False
    line.dataLabels.showSerName = False
    line.legend = Legend()
    line.legend.position = "tr"
    line.legend.overlay = True
    line.legend.layout = Layout(
        manualLayout=ManualLayout(xMode="edge", yMode="edge",
                                  x=0.02, y=0.02, h=0.16, w=0.40))
    ws.add_chart(line, "E24")

    _TIER_ORDER = [
        "1 — Top General Econ",
        "2 — Top Field",
        "3 — Quality Field",
        "4 — Other Peer-Reviewed",
        "WP — Working Paper / Non-Journal",
    ]
    tier_counts = {t: 0 for t in _TIER_ORDER}
    for p in papers:
        jt = p.get("journal_tier") or ""
        if jt in tier_counts:
            tier_counts[jt] += 1
    pie_start = tbl_end + 3
    ws.cell(pie_start, 1, "Journal Tier").font = _F_WHITE_BOLD
    ws.cell(pie_start, 1).fill = _FILL_DARK
    ws.cell(pie_start, 2, "Papers").font = _F_WHITE_BOLD
    ws.cell(pie_start, 2).fill = _FILL_DARK
    pie_labels = [
        "Tier 1 — Top General Econ",
        "Tier 2 — Top Field",
        "Tier 3 — Quality Field",
        "Tier 4 — Other Peer-Reviewed",
        "Working Paper / Non-Journal",
    ]
    for k, (tkey, lbl) in enumerate(zip(_TIER_ORDER, pie_labels), 1):
        ws.cell(pie_start + k, 1, lbl)
        ws.cell(pie_start + k, 2, tier_counts[tkey])
    pie_end = pie_start + len(_TIER_ORDER)

    pie = PieChart()
    pie.title = "Papers by Journal Tiers"
    pie.width, pie.height = 15, 7.5
    pie_data = Reference(ws, min_col=2, min_row=pie_start, max_row=pie_end)
    pie_cats = Reference(ws, min_col=1, min_row=pie_start + 1, max_row=pie_end)
    pie.add_data(pie_data, titles_from_data=True)
    pie.set_categories(pie_cats)
    pie.dataLabels = DataLabelList()
    pie.dataLabels.showPercent = False
    pie.dataLabels.showVal = True
    for idx, clr in enumerate(["1F3864", "2E75B6", "5B9BD5", "9DC3E6", "DEEBF7"]):
        pt = DataPoint(idx=idx)
        pt.graphicalProperties.solidFill = clr
        pie.series[0].data_points.append(pt)
    ws.add_chart(pie, "E47")


def export_excel(papers: list, review: list, search_log: list, output_path: str,
                 min_year: int = 1980, excluded=None):
    if not papers:
        print("[export] No papers to write.")
        return

    def _fy_key(p):
        return (-fy_to_year(p.get("fy") or ""), (p.get("title") or "").lower())

    df = pd.DataFrame(sorted(papers, key=_fy_key))
    for c in OUTPUT_COLS:
        if c not in df.columns:
            df[c] = None
    df = df[[c for c in OUTPUT_COLS if c in df.columns] +
            [c for c in df.columns if c not in OUTPUT_COLS and not c.startswith("_")]]
    try:    df = df.map(clean_cell)
    except: df = df.applymap(clean_cell)

    current_fy, completed_fys = current_and_prior_fy()
    trend_df = _build_trend_df(papers, current_fy)

    _TIER_DESC = {
        "A": "Unambiguous name - full-text match accepted on its own",
        "B": "Generic survey name - needs country context + allowed field",
        "C": "Short acronym - country context + allowed field + case check",
    }
    kw_rows = []
    for f in SURVEY_FAMILIES:
        for term, tier, hints, excl in iter_terms(f):
            kw_rows.append((f["label"], f["region"], term, tier, _TIER_DESC[tier],
                            build_search_query(term, tier, hints, excl),
                            ", ".join(hints) if tier in ("B", "C") else ""))
    kw_df = pd.DataFrame(kw_rows, columns=[
        "Survey Family", "Region", "Search Term", "Match Tier",
        "Matching Rule", "OpenAlex Query Sent", "Required Context Words"])

    PAPERS, ANALYSIS, TREND = "Papers", "Analysis", "FY Trend"
    BACKUP = "Not Relevant (Backup)"
    hdr_fill = PatternFill("solid", start_color="1F5C99", end_color="1F5C99")
    hdr_font = Font(bold=True, color="FFFFFF")

    def _style(ws):
        for cell in ws[1]:
            cell.font = hdr_font
            cell.fill = hdr_fill
            cell.alignment = Alignment(horizontal="center", wrap_text=True)

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name=PAPERS, index=False)
        pd.DataFrame().to_excel(writer, sheet_name=ANALYSIS)
        trend_df.to_excel(writer, sheet_name=TREND, index=False)
        kw_df.to_excel(writer, sheet_name="Keywords", index=False)
        if search_log:
            pd.DataFrame(search_log).to_excel(writer, sheet_name="Search Log", index=False)
        if review:
            pd.DataFrame(review).to_excel(writer, sheet_name="Dedup Review", index=False)

        ex_df = None
        if excluded:
            ex_df = pd.DataFrame(sorted(excluded, key=_fy_key))
            for c in OUTPUT_COLS:
                if c not in ex_df.columns:
                    ex_df[c] = None
            ex_df = ex_df[[c for c in OUTPUT_COLS if c in ex_df.columns]]
            try:    ex_df = ex_df.map(clean_cell)
            except: ex_df = ex_df.applymap(clean_cell)
            ex_df.to_excel(writer, sheet_name=BACKUP, index=False)

        ws_p = writer.sheets[PAPERS]
        ws_p.freeze_panes = "A2"
        _style(ws_p)
        for i, cn in enumerate(df.columns, 1):
            ws_p.column_dimensions[get_column_letter(i)].width = COL_WIDTHS.get(cn, 13)

        if excluded and BACKUP in writer.sheets:
            ws_b = writer.sheets[BACKUP]
            ws_b.freeze_panes = "A2"
            _style(ws_b)
            for i, cn in enumerate(ex_df.columns, 1):
                ws_b.column_dimensions[get_column_letter(i)].width = COL_WIDTHS.get(cn, 13)

        ws_k = writer.sheets["Keywords"]
        ws_k.freeze_panes = "A2"
        _style(ws_k)
        for ltr, w in {"A": 34, "B": 20, "C": 58, "D": 11, "E": 52, "F": 42, "G": 30}.items():
            ws_k.column_dimensions[ltr].width = w

        ws_t = writer.sheets[TREND]
        _style(ws_t)
        for ltr, w in {"A": 8, "B": 14, "C": 20, "D": 20, "E": 16, "F": 16}.items():
            ws_t.column_dimensions[ltr].width = w

    # Analysis sheet is written after the writer closes, so pandas does not
    # interfere with the cell styling.
    wb_out = load_workbook(output_path)
    ws_a = wb_out[ANALYSIS]
    ws_a.delete_rows(1, ws_a.max_row)
    write_analysis_sheet(ws_a, papers, current_fy, completed_fys,
                         min_year, excluded, trend_df)
    wb_out.save(output_path)

    print(f"[export] {len(df)} papers -> {output_path}")
    if excluded:
        print(f"[export] {len(excluded)} no-signal papers -> '{BACKUP}' sheet")
    if review:
        print(f"[export] {len(review)} fuzzy -> 'Dedup Review'")
