"""
discover.py — LSMS Paper Discovery Engine

Searches OpenAlex for papers using LSMS-ISA survey data, filters and scores
them, and writes the results to a formatted Excel workbook.

USAGE:
    python discover.py --api-key YOUR_FREE_KEY                  # full run (~5 min)
    python discover.py --api-key YOUR_FREE_KEY --test           # quick test (30 s)
    python discover.py --api-key YOUR_FREE_KEY --since-fy FY25  # only FY25 onward
    python discover.py --api-key YOUR_FREE_KEY \\
        --merge-existing LSMS_papers_20260101.xlsx             # dedup against existing

DEPENDENCIES:
    pip install requests openpyxl pandas tqdm
"""

import argparse
import re
import time
import unicodedata
import difflib
import hashlib
from datetime import date
from pathlib import Path
from typing import Optional

import html as html_module
import requests
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Alignment

try:
    from tqdm import tqdm
    HAS_TQDM = True
except ImportError:
    HAS_TQDM = False
    def tqdm(x, **kw): return x

from keywords import SURVEY_FAMILIES, SSA_COUNTRY_CODES, SSA_COUNTRY_NAMES, AFRICA_COUNTRY_CODES

# ─────────────────────────────────────────────────────────────────────────────
# CONSTANTS
# ─────────────────────────────────────────────────────────────────────────────

OPENALEX_BASE    = "https://api.openalex.org"
CROSSREF_BASE    = "https://api.crossref.org"

PAGE_SIZE        = 200       # OpenAlex max per-page
BATCH_SIZE       = 8         # terms per OR query (keeps URL < 2048 chars)
MAX_PAGES        = 25       # 25 × 200 = 5000 results/term cap (prevents runaway on broad terms)
RATE_DELAY       = 0.11      # seconds between requests (polite pool ≈ 10 req/s)
CR_MAX_RESULTS   = 100       # Crossref cap when --crossref enabled

# SSA region labels from keywords.py that count as "Africa topic"
AFRICA_REGIONS   = {"Sub-Saharan Africa"}

# ─────────────────────────────────────────────────────────────────────────────
# FISCAL YEAR
# ─────────────────────────────────────────────────────────────────────────────

def fiscal_year(year: int, month: int) -> str:
    """WBG FY: FY26 = Jul 2025–Jun 2026. Month is required; no fallback."""
    return f"FY{str(year + (1 if month >= 7 else 0))[-2:]}"


def fy_start_date(fy_label: str) -> date:
    end_year = 2000 + int(fy_label[2:])
    return date(end_year - 1, 7, 1)


def current_and_prior_fy(today: Optional[date] = None):
    """Return (current_running_fy, list_of_5_recently_completed_fys)."""
    today = today or date.today()
    running_end = today.year + 1 if today.month >= 7 else today.year
    running = f"FY{str(running_end)[-2:]}"
    completed = [f"FY{str(running_end - i)[-2:]}" for i in range(1, 6)]  # 5 most recent completed
    return running, completed


def _fy_to_year(fy: str) -> int:
    """
    'FY27' -> 2027, 'FY99' -> 1999, 'FY00' -> 2000.
    Two-digit years < 70 are 2000s; >= 70 are 1900s.
    Returns 0 when the label is missing/unparseable (sorts last).
    """
    if not fy or not str(fy).startswith("FY"):
        return 0
    try:
        n = int(str(fy)[2:])
    except ValueError:
        return 0
    return 2000 + n if n < 70 else 1900 + n


# ─────────────────────────────────────────────────────────────────────────────
# NORMALISATION (dedup)
# ─────────────────────────────────────────────────────────────────────────────

def norm_title(raw) -> str:
    if not raw:
        return ""
    s = unicodedata.normalize("NFKD", str(raw)).encode("ascii", "ignore").decode()
    s = re.sub(r"[^a-z0-9 ]", " ", s.lower())
    return re.sub(r"\s+", " ", s).strip()


def norm_doi(raw) -> str:
    if not raw:
        return ""
    s = str(raw).lower().strip()
    return re.sub(r"^https?://(dx\.)?doi\.org/", "", s).rstrip("/")

# ─────────────────────────────────────────────────────────────────────────────
# GEOGRAPHY
# ─────────────────────────────────────────────────────────────────────────────

def classify_geography(country_codes: list, first_author_codes: list) -> dict:
    """
    Returns Africa flags based on institutional affiliation.

    IMPORTANT LIMITATION: This approach undercounts African researchers because:
      (1) African scholars at non-African institutions (World Bank, US/EU universities)
          are classified as non-African.
      (2) OpenAlex institution data is incomplete for many African universities
          (papers from African journals often have no institution → 'Unclassified').
    The original WBG tracker (~35%) was manually coded and is more accurate;
    automated institution-based detection is a structural undercount.

    Flags returned:
      geography_clean              : 'Sub-Saharan Africa' | 'Mixed' | 'Other' | 'Unclassified'
      is_first_author_africa       : first author's institution(s) include an African country
      is_any_author_africa         : ANY author has an African institution
      is_africa_institution_strict : ALL institutions are in SSA (narrowest)
    """
    codes    = {c.upper() for c in country_codes      if c}
    fa_codes = {c.upper() for c in first_author_codes if c}

    is_fa_africa  = bool(fa_codes & AFRICA_COUNTRY_CODES)
    is_any_africa = bool(codes    & AFRICA_COUNTRY_CODES)

    if not codes:
        geo    = "Unclassified"
        strict = False
    elif codes <= SSA_COUNTRY_CODES:
        geo    = "Sub-Saharan Africa"
        strict = True
    elif codes & AFRICA_COUNTRY_CODES:
        geo    = "Mixed"
        strict = False
    elif len(codes) > 1:
        geo    = "Mixed"
        strict = False
    else:
        geo    = "Other"
        strict = False

    return {
        "geography_clean":              geo,
        "is_first_author_africa":       is_fa_africa,
        "is_any_author_africa":         is_any_africa,
        "is_africa_institution_strict": strict,
    }

# ─────────────────────────────────────────────────────────────────────────────
# OPENALEX FETCHER  (primary source)
# ─────────────────────────────────────────────────────────────────────────────

_OA_SELECT = ",".join([
    "id","doi","display_name","publication_date","publication_year",
    "type","open_access","authorships","primary_location",
    "cited_by_count","abstract_inverted_index","language","primary_topic",
])


class OpenAlexFetcher:
    def __init__(self, api_key: str = ""):
        self.api_key = api_key
        self.s = requests.Session()
        self.s.headers["User-Agent"] = "LSMS-Tracker/2.0"

    def _get(self, params: dict) -> dict:
        if self.api_key:
            params["api_key"] = self.api_key
        for attempt in range(3):
            try:
                time.sleep(RATE_DELAY)
                r = self.s.get(f"{OPENALEX_BASE}/works", params=params, timeout=60)
                if r.status_code == 429:
                    print("  [rate limit] 429 — waiting 30s. Add --api-key for free 10× budget.", flush=True)
                    time.sleep(30)
                    continue
                r.raise_for_status()
                return r.json()
            except Exception as e:
                if attempt == 2:
                    print(f"  [warn] OpenAlex error: {e}")
                    return {}
                time.sleep(5 * (attempt + 1))
        return {}

    @staticmethod
    def _abstract(inv) -> str:
        if not inv:
            return ""
        return " ".join(w for _, w in sorted(
            ((p, w) for w, ps in inv.items() for p in ps)
        ))

    @staticmethod
    def _authors(authorships: list) -> dict:
        names, countries, insts = [], [], []
        fa_countries = []   # first-author institution countries only
        for idx, a in enumerate(authorships):
            n = (a.get("author") or {}).get("display_name", "")
            if n:
                names.append(n)
            inst_codes = []
            inst_names = []
            for inst in a.get("institutions", []):
                c = inst.get("country_code", "")
                if c:
                    countries.append(c)
                    inst_codes.append(c)
                iname = inst.get("display_name", "")
                if iname:
                    insts.append(iname)
                    inst_names.append(iname)
            if idx == 0:                        # first author only
                fa_countries = inst_codes
        return {
            "authors":               "; ".join(names),
            "first_author":          names[0] if names else "",
            "second_author":         names[1] if len(names) > 1 else "",
            "n_authors":             len(names),
            "affiliations":          "; ".join(dict.fromkeys(insts)),
            "university_affiliation":"; ".join(dict.fromkeys(insts)),  # same data, named to match original
            "affiliation_countries": "; ".join(dict.fromkeys(countries)),
            "_codes":                list(dict.fromkeys(countries)),
            "_fa_codes":             list(dict.fromkeys(fa_countries)),
        }

    def _parse(self, w: dict, family_label: str, family_region: str,
               terms_matched: list) -> dict:
        doi = w.get("doi") or ""
        pd_raw = w.get("publication_date") or ""
        year = month = None
        parts = pd_raw.split("-") if pd_raw else []
        try:
            year  = int(parts[0]) if parts else w.get("publication_year")
            month = int(parts[1]) if len(parts) > 1 else None
        except (ValueError, IndexError):
            year = w.get("publication_year")

        fy = fiscal_year(year, month) if (year and month) else None

        ainfo    = self._authors(w.get("authorships", []))
        codes    = ainfo.pop("_codes", [])
        fa_codes = ainfo.pop("_fa_codes", [])
        geo      = classify_geography(codes, fa_codes)
        loc = w.get("primary_location") or {}
        src = (loc.get("source") or {})
        peer = w.get("type") == "article" or loc.get("is_published", False)

        oa = w.get("open_access") or {}

        abstract_text = self._abstract(w.get("abstract_inverted_index"))
        title_text    = w.get("display_name") or ""
        authorships   = w.get("authorships", [])
        venue_name    = (loc.get("source") or {}).get("display_name", "")
        pub_type_det  = _detect_pub_type(w.get("type", ""), venue_name)
        journal_tier  = _detect_journal_tier(venue_name, pub_type_det)
        topics_auto   = _auto_topics(title_text, abstract_text)
        countries_auto = _dataset_countries(family_label, "; ".join(terms_matched))

        result = {
            "title":               title_text,
            "doi":                 doi,
            "year":                year,
            "month":               month,
            "pub_date":            pd_raw,
            "fy":                  fy,
            # ── Publication type ──
            "publication_type":    w.get("type", ""),
            "pub_type":            pub_type_det,
            "journal_tier":        journal_tier,
            "peer_reviewed_auto":  "Yes" if peer else "No",
            "venue":               venue_name,
            # ── Authors ──
            **ainfo,
            # ── Links ──
            "link":                oa.get("oa_url") or doi,
            "open_access":         oa.get("is_oa", False),
            "oa_url":              oa.get("oa_url") or "",
            # ── Affiliations (original structure + auto-detection) ──
            "wb_affiliation_auto":       _detect_wb(authorships),
            "multilateral_affiliation":  _detect_multilat(authorships),
            # ── Geography ──
            **geo,
            # ── LSMS coding (manual — most important) ──
            # ── Relevance pre-score ──
            "relevance_score":     0,    # filled below
            "relevance_flags":     "",
            # ── Survey dataset country (single column) ──
            "dataset_country": "; ".join(c for c,v in countries_auto.items() if v=="Yes"),
            # ── Research topics (comma-separated auto-detected topics) ──
            "research_topics": ", ".join(k for k,v in topics_auto.items() if v=="Auto"),
            # ── Content ──
            "abstract":            abstract_text,
            "citation_count":      w.get("cited_by_count", 0),
            "language":            w.get("language") or "",
            # ── Pipeline metadata ──
            "survey_family":       family_label,
            "survey_region":       family_region,
            "survey_terms_matched":"; ".join(terms_matched),
            "openalex_id":         w.get("id", ""),
            "source":              "OpenAlex",
            "date_discovered":     date.today().isoformat(),
        }
        _rscore, _rflags = _relevance_score(
            result.get("title",""), result.get("abstract",""),
            result.get("survey_terms_matched","")
        )
        result["relevance_score"] = _rscore
        result["relevance_flags"] = _rflags
        return result

    def search_family(self, family: dict,
                      since_date: Optional[str] = None,
                      family_index: int = 0,
                      total_families: int = 54,
                      verbose: bool = True) -> list[dict]:
        """
        Search each term in the family individually (not batched OR).
        Prints a full line per term BEFORE the API call so the user always
        sees activity — no silent waiting.
        Deduplicates by OpenAlex ID within the family.
        """
        label  = family["label"]
        region = family["region"]
        terms  = family["terms"]   # list of (term, tier) tuples

        seen_ids: set = set()
        results: list = []

        if verbose:
            print(f"[{family_index}/{total_families}] {label} ({len(terms)} terms)", flush=True)

        for term, term_tier in terms:
            # Print the term BEFORE hitting the API — guarantees immediate output
            if verbose:
                print(f"  >> {term[:60]}", flush=True)

            # Build search query: compound AND terms become boolean; simple → phrase
            params_query: dict = {
                "search":   _build_search_query(term, term_tier),
                "select":   _OA_SELECT,
                "per_page": PAGE_SIZE,
            }
            if since_date:
                params_query["filter"] = f"publication_date.gte:{since_date}"

            cursor = "*"
            pages  = 0
            n_rejected = 0
            term_results: list = []

            while True:
                data = self._get({**params_query, "cursor": cursor})
                if not data:
                    break

                meta  = data.get("meta", {})
                items = data.get("results", [])

                if pages == 0 and verbose:
                    count = meta.get("count", "?")
                    print(f"     {count} results", flush=True)

                if not items:
                    break

                for w in items:
                    oa_id = w.get("id", "")
                    if oa_id and oa_id in seen_ids:
                        continue

                    # ── Tiered quality gate (cheap, before expensive _parse) ──
                    inv = w.get("abstract_inverted_index") or {}
                    if inv:
                        pos_words = [(p, wd) for wd, ps in inv.items() for p in ps]
                        abstract_txt = " ".join(wd for _, wd in sorted(pos_words))
                    else:
                        abstract_txt = ""
                    title_txt = w.get("display_name") or ""

                    ok, tier, reason = _passes_filters(
                        title_txt, abstract_txt, term, term_tier,
                        context_hints = family.get("context_hints") or [],
                        primary_topic = w.get("primary_topic"),
                    )
                    if not ok:
                        n_rejected += 1
                        continue

                    if oa_id:
                        seen_ids.add(oa_id)
                    parsed = self._parse(w, label, region, [term])
                    parsed["match_tier"]   = tier
                    parsed["match_reason"] = reason
                    # No abstract on OpenAlex + Tier A match: don't let a data
                    # gap masquerade as low relevance.
                    if tier == "A" and not (parsed.get("abstract") or "").strip():
                        if (parsed.get("relevance_score") or 0) < 2:
                            parsed["relevance_score"] = 2
                            old_flags = parsed.get("relevance_flags") or ""
                            add = "tierA_match_no_abstract_available"
                            parsed["relevance_flags"] = (
                                add if old_flags in ("", "no_strong_signal")
                                else f"{old_flags},{add}"
                            )
                    term_results.append(parsed)
                cursor = meta.get("next_cursor")
                pages += 1
                if pages > 1 and verbose:
                    print(f"     page {pages} ...", flush=True)
                if not cursor or pages >= MAX_PAGES:
                    break

            if verbose and (term_results or n_rejected):
                msg = f"     kept {len(term_results)}"
                if n_rejected:
                    msg += f"  (filtered out {n_rejected})"
                print(msg, flush=True)
            results.extend(term_results)

        if verbose:
            print(f"  --> {len(results)} total for {label}", flush=True)
        return results


# ─────────────────────────────────────────────────────────────────────────────
# CROSSREF FETCHER  (optional supplement — titles only, small cap)
# ─────────────────────────────────────────────────────────────────────────────

class CrossrefFetcher:
    """
    Only used when --crossref flag is set. Searches title fields only
    (`query.title`) with a hard cap of CR_MAX_RESULTS per term, to avoid
    the flood of false positives that `query=` (full-field search) produces.
    Only runs for terms ≥ 18 characters to skip short acronyms.
    """

    def __init__(self, api_key: str = ""):
        self.s = requests.Session()
        self.s.headers["User-Agent"] = "LSMS-Tracker/2.0"

    def _parse(self, item: dict, family_label: str, family_region: str,
               term: str) -> Optional[dict]:
        title_list = item.get("title", [])
        title = title_list[0] if title_list else ""
        if not title:
            return None

        # Verify the term actually appears in the title (Crossref can be fuzzy)
        if term.lower().rstrip(".") not in title.lower():
            return None

        doi = item.get("DOI", "")
        pub = (item.get("published") or item.get("published-print")
               or item.get("published-online") or {})
        dp    = (pub.get("date-parts") or [[]])[0]
        year  = dp[0] if dp else None
        month = dp[1] if len(dp) > 1 else None
        pub_date = f"{year}-{month:02d}" if (year and month) else (str(year) if year else "")
        fy = fiscal_year(year, month) if (year and month) else None

        authors_raw = item.get("author", [])
        names = [f"{a.get('given','')} {a.get('family','')}".strip()
                 for a in authors_raw]
        venue = (item.get("container-title") or [""])[0]
        peer  = item.get("type") in ("journal-article",)
        geo   = classify_geography([], [])

        return {
            "title": title, "doi": doi,
            "year": year, "month": month, "pub_date": pub_date, "fy": fy,
            "publication_type": item.get("type", ""),
            "venue": venue,
            "published_peer_reviewed": peer,
            "open_access": False, "oa_url": "",
            "citation_count": item.get("is-referenced-by-count", 0),
            "abstract": item.get("abstract", ""),
            "language": "",
            "openalex_id": "",
            "authors": "; ".join(names),
            "first_author": names[0] if names else "",
            "n_authors": len(names),
            "affiliations": "",
            "affiliation_countries": "",
            "survey_family": family_label,
            "survey_region": family_region,
            "survey_terms_matched": term,
            "source": "Crossref",
            "date_discovered": date.today().isoformat(),
            **geo,
        }

    def search_term(self, term: str, family: dict) -> list[dict]:
        if len(term) < 18:      # skip short acronyms — too many false positives
            return []
        label, region = family["label"], family["region"]
        results, offset = [], 0
        while offset < CR_MAX_RESULTS:
            try:
                time.sleep(RATE_DELAY)
                r = self.s.get(f"{CROSSREF_BASE}/works", params={
                    "query.title": term,
                    "rows": min(100, CR_MAX_RESULTS - offset),
                    "offset": offset,
                    "select": "DOI,title,author,published,published-print,"
                               "container-title,type,is-referenced-by-count,abstract",
                }, timeout=20)
                r.raise_for_status()
                items = r.json().get("message", {}).get("items", [])
            except Exception:
                break
            if not items:
                break
            for item in items:
                parsed = self._parse(item, label, region, term)
                if parsed:
                    results.append(parsed)
            offset += len(items)
            if len(items) < 100:
                break
        return results

# ─────────────────────────────────────────────────────────────────────────────
# DEDUPLICATION
# ─────────────────────────────────────────────────────────────────────────────

def deduplicate(papers: list[dict],
                existing_df: Optional[pd.DataFrame] = None,
                fuzzy_threshold: float = 0.88) -> tuple[list, list]:
    """
    1) Within-batch dedup (by OpenAlex ID → DOI → exact title).
       When same paper found by multiple terms, merges survey_terms_matched.
    2) Against existing master (exact DOI / title, then fuzzy title).
    Returns (clean_list, review_list).
    """
    # Within-batch
    by_oaid:  dict[str, dict] = {}
    by_doi:   dict[str, dict] = {}
    by_title: dict[str, dict] = {}
    order: list[dict] = []

    for p in papers:
        oa  = p.get("openalex_id", "")
        doi = norm_doi(p.get("doi", ""))
        ttl = norm_title(p.get("title", ""))

        existing = None
        if oa  and oa  in by_oaid:  existing = by_oaid[oa]
        elif doi and doi in by_doi:  existing = by_doi[doi]
        elif ttl and ttl in by_title: existing = by_title[ttl]

        if existing:
            # Merge survey_terms_matched
            old_terms = set((existing.get("survey_terms_matched") or "").split("; "))
            new_terms = set((p.get("survey_terms_matched") or "").split("; "))
            existing["survey_terms_matched"] = "; ".join(sorted(old_terms | new_terms))
            old_fam = set((existing.get("survey_family") or "").split("; "))
            new_fam = set((p.get("survey_family") or "").split("; "))
            existing["survey_family"] = "; ".join(sorted(old_fam | new_fam))
            # Merge dataset_country too (multi-country papers)
            old_dc = set(x.strip() for x in (existing.get("dataset_country") or "").split(";") if x.strip())
            new_dc = set(x.strip() for x in (p.get("dataset_country") or "").split(";") if x.strip())
            existing["dataset_country"] = "; ".join(sorted(old_dc | new_dc))
            continue

        if oa:  by_oaid[oa]   = p
        if doi: by_doi[doi]   = p
        if ttl: by_title[ttl] = p
        order.append(p)

    # A paper matching 2+ distinct survey families is almost certainly using
    # LSMS microdata even if its abstract only says "survey data" — bump it
    # above the relevance cutoff so it isn't stranded in the backup sheet.
    for p in order:
        fams = [x for x in (p.get("survey_family") or "").split("; ") if x.strip()]
        if len(set(fams)) >= 2 and (p.get("relevance_score") or 0) < 2:
            p["relevance_score"] = 2
            existing_flags = p.get("relevance_flags") or ""
            add = f"multi_survey_match_{len(set(fams))}"
            p["relevance_flags"] = (
                add if existing_flags in ("", "no_strong_signal")
                else f"{existing_flags},{add}"
            )

    # Against existing master
    review: list = []
    if existing_df is None or existing_df.empty:
        return order, review

    ex_dois = set()
    ex_titles: list[str] = []
    for col in ("doi",):
        if col in existing_df.columns:
            ex_dois = {norm_doi(str(v)) for v in existing_df[col].dropna()}
    for col in ("title", "Document Info"):
        if col in existing_df.columns:
            ex_titles = [norm_title(str(v)) for v in existing_df[col].dropna()]
            break

    ex_title_set = set(ex_titles)
    clean: list = []
    for p in order:
        doi = norm_doi(p.get("doi", ""))
        ttl = norm_title(p.get("title", ""))

        if doi and doi in ex_dois:
            continue
        if ttl and ttl in ex_title_set:
            continue
        if ttl and ex_titles:
            best = max(
                (difflib.SequenceMatcher(None, ttl, et).ratio()
                 for et in ex_titles if abs(len(et) - len(ttl)) < 40),
                default=0.0,
            )
            if best >= fuzzy_threshold:
                p["_fuzzy_score"] = round(best, 3)
                review.append(p)
                continue
        clean.append(p)

    return clean, review

# ─────────────────────────────────────────────────────────────────────────────
# EXCEL EXPORT
# ─────────────────────────────────────────────────────────────────────────────

OUTPUT_COLS = [
    "title", "doi", "year", "month", "pub_date", "fy",
    "publication_type", "pub_type", "journal_tier", "peer_reviewed_auto", "venue",
    "authors", "first_author", "second_author", "n_authors",
    "link", "open_access", "oa_url",
    "wb_affiliation_auto",
    "multilateral_affiliation",
    "university_affiliation",
    "affiliations",
    "affiliation_countries",
    "geography_clean",
    "is_first_author_africa",
    "is_any_author_africa",
    "is_africa_institution_strict",
    "relevance_score",
    "relevance_flags",
    "match_tier",
    "match_reason",
    "dataset_country",
    "research_topics",
    "abstract",
    "citation_count",
    "language",
    "survey_family",
    "survey_terms_matched",
    "openalex_id",
    "source",
    "date_discovered",
]

COL_WIDTHS = {
    "title": 60, "doi": 34, "abstract": 60,
    "authors": 35, "affiliations": 40, "affiliation_countries": 20,
    "university_affiliation": 35, "multilateral_affiliation": 25,
    "venue": 25, "survey_family": 30, "survey_terms_matched": 38,
    "relevance_flags": 30, "match_reason": 34, "geography_clean": 20,
}


def _build_trend_df(papers: list, current_fy: str) -> "pd.DataFrame":
    """Build the per-FY trend data used by both the Analysis chart and FY Trend sheet."""
    from keywords import SURVEY_FAMILIES as _KW  # noqa — used elsewhere
    fy_labels = [f"FY{str(y)[-2:]}" for y in range(2009, 2030)]
    rows = []
    for fy in fy_labels:
        fp = [p for p in papers if p.get("fy") == fy]
        rows.append((fy, len(fp),
                     sum(1 for p in fp if p.get("is_any_author_africa")),
                     sum(1 for p in fp if p.get("is_first_author_africa")),
                     sum(1 for p in fp if (p.get("peer_reviewed_auto") or "") == "Yes"),
                     sum(1 for p in fp if (p.get("wb_affiliation_auto") or "") == "Yes")))
    return pd.DataFrame(rows, columns=["FY","All Papers","Africa (any author)",
                                       "Africa (1st author)","Peer-Reviewed","WB-Affiliated"])

# ─────────────────────────────────────────────────────────────────────────────
# SEARCH QUALITY FILTERS — three-tier precision/recall system
#
# The keyword list contains terms of wildly different specificity. One filter
# applied to all of them either loses real papers (too strict on long phrases)
# or admits garbage (too loose on short acronyms). Terms are tiered:
#
#   TIER A — UNAMBIGUOUS: >=3 words, or contains a distinctive survey phrase.
#            e.g. "Uganda National Panel Survey", "Living Standards Measurement Study"
#            Such strings cannot plausibly appear outside LSMS-related work.
#            -> Accept OpenAlex full-text matches. No title/abstract requirement,
#               no discipline filter. This is where most real papers come from:
#               surveys are named in the Data/Methods section, not the abstract.
#
#   TIER B — MEDIUM: 2-word terms or hyphenated/mixed-case identifiers.
#            e.g. "LSMS-ISA", "Tanzania NPS", "GHS-Panel", "IHS3-Panel"
#            -> Require the term in title OR abstract (case-insensitive,
#               word boundary). No discipline filter.
#
#   TIER C — AMBIGUOUS ACRONYM: single ALL-CAPS token <= 6 chars.
#            e.g. "IHPS", "ESS1", "UNPS", "IHS", "HFPS", "TNPS"
#            -> Require ALL of:
#                 (a) case-SENSITIVE word-boundary match
#                     ("IHPS" != "IHPs", plural of the medical term IHP)
#                 (b) at least one country/context hint word present
#                 (c) paper not in an obviously irrelevant discipline
#
#   COMPOUND AND terms ("HFPS and Burkina Faso") split into a boolean AND query;
#   each component is then checked at its own tier.
# ─────────────────────────────────────────────────────────────────────────────

TIER_A = "A"
TIER_B = "B"
TIER_C = "C"

# Disciplines where a household living-standards survey cannot be the data source.
# Deliberately narrow: health, nutrition, agriculture, education all stay IN.
_EXCL_DOMAINS   = {"Physical Sciences"}
_EXCL_FIELDS    = {
    "Arts and Humanities",
    "Biochemistry, Genetics and Molecular Biology",
    "Immunology and Microbiology",
    "Neuroscience",
    "Chemical Engineering",
    "Materials Science",
    "Dentistry",
    "Pharmacology, Toxicology and Pharmaceutics",
}
_EXCL_SUBFIELDS = {
    "Linguistics and Language", "Philosophy",
    "Literature and Literary Theory", "Visual Arts and Performing Arts",
    "Astronomy and Astrophysics", "Particle and Nuclear Physics",
    "Organic Chemistry", "Inorganic Chemistry", "Physical Chemistry",
    "Condensed Matter Physics", "Cell Biology", "Genetics",
    "Film, Television, and Video Studies", "Cultural Studies",
}


def _is_compound_and(tier: str) -> bool:
    """True when this term's tier (as given in keywords.py) is AND."""
    return tier == "AND"


def _build_search_query(term: str, tier: str) -> str:
    """OpenAlex search= string. Compound AND -> boolean; otherwise phrase."""
    if _is_compound_and(tier):
        parts = re.split(r"\s+and\s+", term, flags=re.IGNORECASE)
        return " AND ".join('"' + p.strip() + '"' for p in parts)
    return '"' + term + '"'


def _word_present(word: str, text: str, case_sensitive: bool = False) -> bool:
    """Word-boundary match. 'niger' will not match 'Nigeria'."""
    flags = 0 if case_sensitive else re.IGNORECASE
    return bool(re.search(r"\b" + re.escape(word) + r"\b", text, flags))


def _is_relevant_topic(primary_topic: Optional[dict]) -> bool:
    """Only used for TIER C terms. Permissive when topic data is missing."""
    if not primary_topic:
        return True
    domain   = (primary_topic.get("domain")   or {}).get("display_name", "")
    field    = (primary_topic.get("field")    or {}).get("display_name", "")
    subfield = (primary_topic.get("subfield") or {}).get("display_name", "")
    return (domain   not in _EXCL_DOMAINS and
            field    not in _EXCL_FIELDS   and
            subfield not in _EXCL_SUBFIELDS)


def _passes_filters(title: str, abstract: str, term: str, tier: str,
                    context_hints: Optional[list],
                    primary_topic: Optional[dict]) -> tuple:
    """
    Master gate. `tier` is the classification assigned to this term directly
    in keywords.py — A / B / C / AND. Returns (passes: bool, tier: str, reason: str).
    """
    if tier == "AND":
        parts = re.split(r"\s+and\s+", term, flags=re.IGNORECASE)
        for p in parts:
            # Sub-parts of a compound term (e.g. "IHS3" and "Panel" from
            # "IHS3 and Panel") don't have their own listed tier, so classify
            # them at Tier B (must appear in title/abstract) by default —
            # a plain word-boundary check without the extra Tier C strictness.
            ok, _, why = _passes_filters(title, abstract, p.strip(), "B",
                                         context_hints=None,
                                         primary_topic=primary_topic)
            if not ok:
                return False, "AND", "component '" + p.strip() + "': " + why
        return True, "AND", "all AND components matched"

    text = title + " " + abstract

    if tier == "A":
        return True, "A", "unambiguous term (fulltext match accepted)"

    if tier == "B":
        if _word_present(term, text, case_sensitive=False):
            return True, "B", "term in title/abstract"
        return False, "B", "term not in title/abstract"

    # TIER C
    if not _word_present(term, text, case_sensitive=True):
        return False, "C", "acronym absent (case-sensitive check)"
    if context_hints and not any(_word_present(h, text, case_sensitive=False)
                                 for h in context_hints):
        return False, "C", "no country/context word near acronym"
    if not _is_relevant_topic(primary_topic):
        return False, "C", "irrelevant discipline"
    return True, "C", "acronym + context + relevant discipline"


# ─────────────────────────────────────────────────────────────────────────────
# RELEVANCE SCORING  (how likely does this paper USE LSMS data vs merely cite it)
# ─────────────────────────────────────────────────────────────────────────────

_USE_PATTERNS    = ["using data from","we use","we analyze","we analyse","we employ",
                    "data come from","data are from","data is from","data drawn from",
                    "draw on data","this paper uses","this study uses","this study analyzes",
                    "this paper analyzes","this paper analyses","using the survey",
                    "household-level data","microdata","panel data from","survey data from",
                    "dataset from","we exploit","primary data","nationally representative",
                    "drawing on data","drawing on the","drawn from the","based on data",
                    "based on the survey","relying on data","relies on data","data collected",
                    "data collected by","collected as part of","this data comes from",
                    "utilizing data","utilising data",]
_REVIEW_PATTERNS = [
    # Only TRUE reviews/meta-analyses — NOT viewpoints, overviews, or practice guides
    "systematic review", "meta-analysis", "scoping review", "bibliometric",
    "narrative review", "rapid review",
    "we conducted a review", "this review article",
    "literature review of", "review of the literature",
    # Definitively NOT empirical papers:
    "this paper reviews the", "we survey the literature",
]
# NOTE: "overview of" and generic "this paper reviews" removed — too aggressive.
# A paper that "provides an overview of [LSMS survey results]" is still an LSMS paper.
_EMPIRICAL_SIGNALS = ["regression","estimate","coefficient","ols ","2sls","probit","logit",
                      "instrumental variable","household","consumption","expenditure",
                      "welfare","poverty","income","fixed effect","random effect",]

# Words that, found near a survey-name mention, indicate the paper actually
# uses that survey's data — looser than _USE_PATTERNS since it only has to
# fire within a tight window around a term already known to be present.
_PROXIMITY_DATA_WORDS = [
    "data", "survey", "dataset", "sample", "collected", "fielded",
    "administered", "drawing", "drawn", "based", "utilized", "utilised",
    "used", "using", "employ", "employed", "wave", "panel", "microdata",
    "respondents", "households", "conducted", "representative",
]
_PROXIMITY_WINDOW_CHARS = 175   # characters to each side of the fuzzy match


def _fuzzy_survey_pattern(term: str):
    """
    Regex for `term` that tolerates adjectival country forms in running
    prose ("Ethiopia" -> "Ethiopian", "Uganda" -> "Ugandan", etc.) by
    matching each word as a prefix rather than requiring an exact match.
    """
    words = [w for w in re.split(r"\s+", term.strip()) if w]
    if not words:
        return None
    parts = [r"\b" + re.escape(w) + r"\w*" for w in words]
    pattern = r"\s+".join(parts)
    try:
        return re.compile(pattern, re.IGNORECASE)
    except re.error:
        return None


def _proximity_data_use_match(terms: list, abstract_original_case: str) -> bool:
    """
    Fuzzy-match each survey term in the abstract; if found, check ~175
    characters either side for data-use vocabulary. Catches cases where the
    abstract uses a grammatical variant of the matched term (e.g. "Ethiopian
    Rural Socioeconomic Survey" vs. the keyword "Ethiopia Rural Socioeconomic
    Survey") that an exact substring check would miss.
    """
    if not abstract_original_case:
        return False
    for term in terms:
        if len(term.strip()) <= 4:
            continue
        pattern = _fuzzy_survey_pattern(term)
        if not pattern:
            continue
        m = pattern.search(abstract_original_case)
        if not m:
            continue
        start = max(0, m.start() - _PROXIMITY_WINDOW_CHARS)
        end   = min(len(abstract_original_case), m.end() + _PROXIMITY_WINDOW_CHARS)
        window = abstract_original_case[start:end].lower()
        if any(w in window for w in _PROXIMITY_DATA_WORDS):
            return True
    return False


def _relevance_score(title: str, abstract: str, survey_terms_matched: str) -> tuple:
    """
    Returns (score: int 0-3, flags: str).

    Score 3: survey term in TITLE                       -> near-certain data use
    Score 2: survey term in ABSTRACT, or explicit
             data-use language, or strong empirical
             signals                                    -> very likely data use
    Score 1: some evidence, but weak
    Score 0: review/meta-analysis detected

    flags == "no_strong_signal" means: the search term never appeared in the
    title or abstract, no data-use language, no empirical vocabulary, and no
    LSMS/World Bank mention. Only OpenAlex's full-text index connected this
    paper to the keyword. These are the papers routed to the backup sheet.
    """
    t     = (title or "").lower()
    a     = (abstract or "").lower()
    full  = t + " " + a
    terms = [s.strip().lower() for s in (survey_terms_matched or "").split(";")
             if len(s.strip()) > 4]

    score = 1
    flags = []

    # (1) Survey term appears in the TITLE — strongest possible signal
    if any(term in t for term in terms):
        score = 3
        flags.append("survey_in_title")

    # (2) Survey term appears in the ABSTRACT — very strong signal
    elif any(term in a for term in terms):
        score = max(score, 2)
        flags.append("survey_in_abstract")

    # (3) Explicit data-use language in the abstract
    if any(p in a for p in _USE_PATTERNS):
        score = max(score, 2)
        flags.append("use_language_in_abstract")

    # (3b) Fuzzy survey match near data-use vocabulary in the abstract
    if "survey_in_abstract" not in flags and "use_language_in_abstract" not in flags:
        if _proximity_data_use_match(terms, abstract or ""):
            score = max(score, 2)
            flags.append("proximity_data_use_match")

    # (4) The programme itself is named (LSMS / World Bank survey work)
    if any(p in full for p in ("lsms", "living standards measurement",
                               "world bank")):
        score = max(score, 2)
        flags.append("lsms_or_worldbank_named")

    # (5) Empirical-paper vocabulary
    emp = sum(1 for e in _EMPIRICAL_SIGNALS if e in full)
    if emp >= 3:
        score = max(score, 2)
        flags.append(f"empirical_signals_{emp}")

    # (6) Review / meta-analysis penalty
    if any(r in full for r in _REVIEW_PATTERNS):
        score = max(0, score - 2)
        flags.append("review_or_meta_analysis")

    return max(0, min(3, score)), (",".join(flags) if flags else "no_strong_signal")


# ─────────────────────────────────────────────────────────────────────────────
# AFFILIATION DETECTION (WB, multilaterals, auto-topic hints)
# ─────────────────────────────────────────────────────────────────────────────

_WB_NAMES = {
    "world bank", "world bank group", "international finance corporation",
    "ibrd", "ida ", "international development association",
}
_MULTILAT_MAP = {
    "IFPRI":  ["international food policy research", "ifpri"],
    "FAO":    ["food and agriculture organization", "fao"],
    "IFAD":   ["international fund for agricultural development", "ifad"],
    "CGIAR":  ["cgiar", "cimmyt", "cip ", "ilri", "cifor", "bioversity",
               "worldfish", "iita", "icarda", "icrisat", "iwmi"],
    "WFP":    ["world food programme", "wfp"],
    "UNICEF": ["unicef"],
    "WHO":    ["world health organization", "who "],
    "UNDP":   ["undp", "united nations development programme"],
    "IMF":    ["international monetary fund", "imf"],
    "AfDB":   ["african development bank", "afdb"],
    "UN":     ["united nations"],
}
_TOPIC_KEYWORDS = {
    "Agricultural Prod":        ["crop", "yield", "harvest", "farm production", "agricultural output",
                                 "maize", "rice", "wheat", "sorghum", "cassava", "groundnut", "livestock"],
    "Agricultural Inputs":      ["fertilizer", "seed", "input", "pesticide", "irrigation", "extension"],
    "Market & Value Chain":     ["market", "value chain", "price transmission", "trade"],
    "Env. & Climate Change":    ["climate", "rainfall", "drought", "temperature", "flood",
                                 "environment", "deforestation"],
    "Other Livelihoods":        ["livelihood", "off-farm", "diversification", "non-farm"],
    "Labor & Time Use":         ["labor", "labour", "employment", "wage", "work", "time use"],
    "Finance":                  ["credit", "loan", "microfinance", "savings", "financial inclusion",
                                 "remittance", "mobile money"],
    "Health":                   ["health", "mortality", "disease", "malaria", "hiv", "stunting",
                                 "maternal", "child health", "morbidity"],
    "Nutrition & Food Security":["nutrition", "food security", "hunger", "dietary", "malnutrition",
                                 "wasting", "anaemia", "anemia"],
    "Gender":                   ["gender", "women", "female", "girl", "empowerment", "intrahousehold"],
    "Poverty, Income, & Welfare":["poverty", "welfare", "consumption", "expenditure", "income",
                                  "living standard", "inequality", "wealth"],
    "Education & Training":     ["education", "school", "learning", "literacy", "dropout", "enrolment"],
}
# Which country each survey family maps to (for auto-filling dataset country columns)
_FAMILY_COUNTRY = {
    "Burkina Faso EMC / EHCVM": "Burkina Faso",
    "Ethiopia ESS / ESPS":      "Ethiopia",
    "Malawi IHS / IHPS":        "Malawi",
    "Mali EACI":                "Mali",
    "Niger ECVMA":              "Niger",
    "Nigeria GHS-Panel":        "Nigeria",
    "Tanzania NPS":             "Tanzania",
    "Uganda UNPS":              "Uganda",
    # HFPS — country extracted from compound AND term
}
_HFPS_COUNTRY_TERMS = {
    "Burkina Faso": "burkina", "Mali": "mali", "Nigeria": "nigeria",
    "Niger": "niger", "Ethiopia": "ethiopia", "Uganda": "uganda",
    "Malawi": "malawi", "Tanzania": "tanzania",
}
_ISA_COUNTRIES = ["Burkina Faso","Ethiopia","Malawi","Mali","Niger","Nigeria","Tanzania","Uganda"]


def _detect_wb(authorships: list) -> str:
    for a in authorships:
        for inst in a.get("institutions", []):
            name = inst.get("display_name","").lower()
            if any(wb in name for wb in _WB_NAMES):
                return "Yes"
    return "No"


def _detect_multilat(authorships: list) -> str:
    found = []
    for org_label, patterns in _MULTILAT_MAP.items():
        for a in authorships:
            for inst in a.get("institutions", []):
                name = inst.get("display_name","").lower()
                if any(p in name for p in patterns):
                    if org_label not in found:
                        found.append(org_label)
    return "; ".join(found) if found else ""


def _auto_topics(title: str, abstract: str) -> dict:
    """Return topic → 'Auto' when keywords from abstract suggest it, else ''."""
    text = (title + " " + abstract).lower()
    out = {}
    for topic, kws in _TOPIC_KEYWORDS.items():
        out[topic] = "Auto" if any(k in text for k in kws) else ""
    return out


def _dataset_countries(survey_family: str, survey_terms_matched: str) -> dict:
    """Return {country: 'Yes'} based on which survey family matched the paper."""
    out = {c: "" for c in _ISA_COUNTRIES}
    # Direct family → country
    country = _FAMILY_COUNTRY.get(survey_family)
    if country:
        out[country] = "Yes"
    # HFPS compound AND terms carry the country in the term name
    if "HFPS" in survey_family or "High-Frequency" in survey_family:
        for term in (survey_terms_matched or "").split(";"):
            term_l = term.strip().lower()
            for country, hint in _HFPS_COUNTRY_TERMS.items():
                if hint in term_l:
                    out[country] = "Yes"
    # Handle multi-family papers (survey_family can be "; ".join(multiple))
    for fam in survey_family.split("; "):
        c2 = _FAMILY_COUNTRY.get(fam.strip())
        if c2:
            out[c2] = "Yes"
    return out


# ─────────────────────────────────────────────────────────────────────────────
# JOURNAL TYPE AND TIER CLASSIFICATION
# ─────────────────────────────────────────────────────────────────────────────

def _detect_pub_type(oa_type: str, venue: str) -> str:
    """
    Classify publication type. Recognises repositories, eBooks, thesis archives,
    and working-paper series as non-journal outputs (these are the OpenAlex
    'venues' that are not actually peer-reviewed journals).
    """
    v = (venue or "").lower()
    t = (oa_type or "").lower()

    # ── Non-journal venue patterns (repositories, theses, eBooks, WP series) ──
    _REPO_MARKERS = (
        "repository", "repositories", "archive", "eprints", "escholarship",
        "scholarship", "research online", "research commons", "open research",
        "digital commons", "academic commons", "brage", "dash", "dspace",
        "udspace", "tspace", "vtechworks", "deep blue", "doria", "duo research",
        "epsilon", "figshare", "mendeley data", "dataverse", "harvard dataverse",
        "icpsr", "swedish national data", "networked services", "dans",
        "institutional repository", "knowledge repository", "open knowledge",
        "research archive", "publications explorer", "research portal",
        "research product catalog", "researchspace", "scholarspace",
        "open repository", "doctoral", "theses", "thesis", "dissertation",
        "etd", "phd", "e-zone", "runde", "macau", "refubium", "bibsys",
        "agecon search", "agecon", "core (", "semantic scholar", "openalex",
        "zenodo", "preprints.org", "research square", "ssoar", "munich personal repec",
        "mpra", "ageconsearch", "knowledge for policy", "opendata",
    )
    _EBOOK_MARKERS = (
        "ebooks", "ebook", "university press", "press ebooks", "elsevier",
        "springer", "palgrave", "edward elgar", "mit press", "intechopen",
        "cambridge university press", "oxford university press", "sage encyclopedia",
        "lambert academic", "ecorfan", "practical action publishing",
        "in-house reproduction", "umi ebooks", "united nations ebooks",
        "world bank ebooks", "fao ebooks", "ilo ebooks", "cifor ebooks",
        "world bank publications", "world bank open knowledge",
    )
    _WP_MARKERS = (
        "working paper", "discussion paper", "technical paper",
        "national bureau of economic research", "nber", "repec",
        "research papers in economics", "econstor", "ssrn", "cepr", "iza",
        "policy research working paper", "opengrey", "opendocs", "hal",
        "open science framework", "osf", "preprint", "arxiv", "philpapers",
        "contributions to economics", "research series", "series",
    )

    if t == "book-chapter" or any(x in v for x in _EBOOK_MARKERS):
        return "Book / eBook"
    if t == "report":
        return "Report"
    if t == "dissertation" or any(x in v for x in ("theses","thesis","dissertation","doctoral","etd")):
        return "Thesis / Dissertation"
    if t == "preprint" or "arxiv" in v or "preprint" in v:
        return "Preprint / Working Paper"
    if any(x in v for x in _WP_MARKERS):
        return "Working Paper"
    if any(x in v for x in _REPO_MARKERS):
        return "Repository / Working Paper"
    if t == "article":
        return "Journal Article"
    if t in ("proceedings-article", "conference-paper"):
        return "Conference Paper"
    if t in ("editorial", "letter"):
        return "Editorial / Letter"
    return "Other"


# Tier lookup — based on standard field rankings for development economics.
# Tiers will be updated with WBG-provided list once available.
# Format: {normalized journal name: tier label}
_JOURNAL_TIERS = {
    # ── Tier 1: Top-5 general economics ───────────────────────────────────────
    "american economic review": "1 — Top General Econ",
    "quarterly journal of economics": "1 — Top General Econ",
    "journal of political economy": "1 — Top General Econ",
    "review of economic studies": "1 — Top General Econ",
    "econometrica": "1 — Top General Econ",
    # ── Tier 2: Top field journals (development / applied) ────────────────────
    "journal of development economics": "2 — Top Field",
    "world development": "2 — Top Field",
    "american economic journal applied economics": "2 — Top Field",
    "american economic journal economic policy": "2 — Top Field",
    "review of economics and statistics": "2 — Top Field",
    "economic journal": "2 — Top Field",
    "journal of economic growth": "2 — Top Field",
    "economic development and cultural change": "2 — Top Field",
    "world bank economic review": "2 — Top Field",
    "journal of the european economic association": "2 — Top Field",
    "rand journal of economics": "2 — Top Field",
    "journal of public economics": "2 — Top Field",
    "journal of international economics": "2 — Top Field",
    "journal of human resources": "2 — Top Field",
    "journal of economic perspectives": "2 — Top Field",
    "journal of urban economics": "2 — Top Field",
    # ── Tier 3: Quality field journals ────────────────────────────────────────
    "food policy": "3 — Quality Field",
    "journal of agricultural economics": "3 — Quality Field",
    "agricultural economics": "3 — Quality Field",
    "american journal of agricultural economics": "3 — Quality Field",
    "journal of african economies": "3 — Quality Field",
    "african development review": "3 — Quality Field",
    "oxford development studies": "3 — Quality Field",
    "health economics": "3 — Quality Field",
    "journal of health economics": "3 — Quality Field",
    "demography": "3 — Quality Field",
    "land economics": "3 — Quality Field",
    "european review of agricultural economics": "3 — Quality Field",
    "world bank research observer": "3 — Quality Field",
    "journal of rural studies": "3 — Quality Field",
    "global food security": "3 — Quality Field",
    "applied economic perspectives and policy": "3 — Quality Field",
    "economics of education review": "3 — Quality Field",
    "labour economics": "3 — Quality Field",
    "journal of development studies": "3 — Quality Field",
    "canadian journal of development studies": "3 — Quality Field",
    "journal of african economies": "3 — Quality Field",
    "international food and agribusiness management review": "3 — Quality Field",
    # ── Tier 3 additions (recognised field journals from corpus) ──────────────
    "food security": "3 — Quality Field",
    "agriculture & food security": "3 — Quality Field",
    "public health nutrition": "3 — Quality Field",
    "journal of health population and nutrition": "3 — Quality Field",
    "health policy and planning": "3 — Quality Field",
    "educational research review": "3 — Quality Field",
    "economics of education review": "3 — Quality Field",
    "world development perspectives": "3 — Quality Field",
    "the european journal of development research": "3 — Quality Field",
    "review of development economics": "3 — Quality Field",
    "journal of international development": "3 — Quality Field",
    "agricultural and food economics": "3 — Quality Field",
    "food and energy security": "3 — Quality Field",
    "population and development review": "3 — Quality Field",
    "social science & medicine": "3 — Quality Field",
    "american journal of epidemiology": "3 — Quality Field",
    "maternal & child nutrition": "3 — Quality Field",
    "journal of nutrition": "3 — Quality Field",
    # ── Tier 4: Other peer-reviewed (recognised, non-top journals) ────────────
    "plos one": "4 — Other Peer-Reviewed",
    "plos global public health": "4 — Other Peer-Reviewed",
    "sustainability": "4 — Other Peer-Reviewed",
    "nutrients": "4 — Other Peer-Reviewed",
    "heliyon": "4 — Other Peer-Reviewed",
    "scientific reports": "4 — Other Peer-Reviewed",
    "bmc public health": "4 — Other Peer-Reviewed",
    "bmc global and public health": "4 — Other Peer-Reviewed",
    "frontiers in public health": "4 — Other Peer-Reviewed",
    "frontiers in health services": "4 — Other Peer-Reviewed",
    "annals of global health": "4 — Other Peer-Reviewed",
    "aids care": "4 — Other Peer-Reviewed",
    "obesity reviews": "4 — Other Peer-Reviewed",
    "springerplus": "4 — Other Peer-Reviewed",
    "people and nature": "4 — Other Peer-Reviewed",
    "emerging infectious diseases": "4 — Other Peer-Reviewed",
    "revista panamericana de salud publica": "4 — Other Peer-Reviewed",
    "salud publica de mexico": "4 — Other Peer-Reviewed",
    "journal of korean medical science": "4 — Other Peer-Reviewed",
    "journal of nepal health research council": "4 — Other Peer-Reviewed",
    "journal of water sanitation and hygiene for development": "4 — Other Peer-Reviewed",
    "statistical journal of the iaos": "4 — Other Peer-Reviewed",
    "statistics in transition new series": "4 — Other Peer-Reviewed",
    "international journal of environmental research and public health": "4 — Other Peer-Reviewed",
    # ── Working papers ────────────────────────────────────────────────────────
    "nber": "WP — Working Paper",
    "iza": "WP — Working Paper",
    "ssrn": "WP — Working Paper",
    "cepr": "WP — Working Paper",
    "world bank policy research": "WP — Working Paper",
    "ifpri discussion": "WP — Working Paper",
}


def _detect_journal_tier(venue: str, pub_type: str) -> str:
    """
    Assign a journal tier.

    Rules (in order):
      1. Working papers / repositories / theses / preprints / eBooks -> 'WP'
      2. Exact or substring match in the curated tier list -> that tier
      3. Any remaining journal article -> Tier 4 (never left blank)
      4. Anything else with no venue -> blank
    """
    pt = pub_type or ""

    # Non-journal outputs -> Working Paper bucket
    if any(x in pt for x in ("Working Paper", "Repository", "Preprint",
                             "Thesis", "Dissertation", "Report",
                             "Book", "eBook", "Conference")):
        return "WP — Working Paper / Non-Journal"

    v = (venue or "").lower().strip()
    if v:
        if v in _JOURNAL_TIERS:
            return _JOURNAL_TIERS[v]
        for key, tier in _JOURNAL_TIERS.items():
            if key in v:
                return tier

    # Recognised journal article but unlisted journal -> lowest tier, never blank
    if "Journal Article" in pt:
        return "4 — Other Peer-Reviewed"

    # Editorials/letters or truly unknown venue
    if v:
        return "4 — Other Peer-Reviewed"
    return ""

# ─────────────────────────────────────────────────────────────────────────────
# ANALYSIS SHEET WRITER
# ─────────────────────────────────────────────────────────────────────────────
# ANALYSIS SHEET
# Cell-by-cell layout matching the reference file: navy headers, light-blue
# highlight cells, 75/48/23/23 column widths. Charts anchored at column E.
# ─────────────────────────────────────────────────────────────────────────────

from openpyxl.styles import Border, Side
from openpyxl.chart import BarChart, LineChart, PieChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.legend import Legend
from openpyxl.chart.layout import Layout, ManualLayout
from openpyxl.chart.marker import DataPoint

# ── Brand colours ──────────────────────────────────────────────────────────
C_DARK  = "1F5C99"   # section header fill (navy blue)
C_HL    = "C5D9F1"   # highlighted count/share cell fill (light blue)
C_WHITE = "FFFFFF"

_THIN_BLACK = Side(style="thin", color="000000")
_BORDER_ALL = Border(left=_THIN_BLACK, right=_THIN_BLACK,
                     top=_THIN_BLACK, bottom=_THIN_BLACK)

_F_WHITE_BOLD = Font(bold=True, color=C_WHITE)
_F_BOLD       = Font(bold=True)
_F_BOLD_BLACK = Font(bold=True, color="000000")
_FILL_DARK    = PatternFill("solid", start_color=C_DARK, end_color=C_DARK)
_FILL_HL      = PatternFill("solid", start_color=C_HL,   end_color=C_HL)
_AL_CENTER    = Alignment(horizontal="center")
_AL_LEFT      = Alignment(horizontal="left")


def _cell(ws, r, c, value=None, bold=False, white_text=False, fill=None, pct=False):
    cell = ws.cell(r, c)
    if value is not None:
        cell.value = value
    if bold or white_text:
        cell.font = Font(bold=bold, color=C_WHITE if white_text else "000000")
    if fill:
        cell.fill = fill
    if pct and isinstance(value, (int, float)):
        cell.number_format = "0%"
    return cell


def _section_row(ws, r, label):
    """Full-width dark-blue section header spanning all 4 columns, black border."""
    c = ws.cell(r, 1, label)
    c.font = _F_WHITE_BOLD
    c.fill = _FILL_DARK
    c.border = _BORDER_ALL
    for col in range(2, 5):
        cc = ws.cell(r, col)
        cc.fill = _FILL_DARK
        cc.border = _BORDER_ALL
    ws.merge_cells(f"A{r}:D{r}")


def _highlight(cell, pct=False):
    """Apply the light-blue highlight: fill C5D9F1, bold black text, black border."""
    cell.fill = _FILL_HL
    cell.font = _F_BOLD_BLACK
    cell.border = _BORDER_ALL
    if pct:
        cell.number_format = "0%"


def _data_row(ws, r, label, note="", count=None, share=None, hl_count=False, hl_share=False, bold_label=False):
    lc = ws.cell(r, 1, label)
    if bold_label:
        lc.font = _F_BOLD
    ws.cell(r, 2, note)
    if count is not None:
        cc = ws.cell(r, 3, count)
        if hl_count:
            _highlight(cc)
        elif bold_label:
            cc.font = _F_BOLD
    if share is not None:
        sc = ws.cell(r, 4, share)
        sc.number_format = "0%"
        if hl_share:
            _highlight(sc, pct=True)


def write_analysis_sheet(ws, papers, current_fy, completed_fys,
                         min_year, excluded, trend_df):
    excluded = excluded or []
    total  = len(papers)
    n0     = sum(1 for p in excluded if (p.get("relevance_score") or 0) == 0)
    n1     = sum(1 for p in excluded if (p.get("relevance_score") or 0) == 1)
    peer   = sum(1 for p in papers if (p.get("peer_reviewed_auto") or "") == "Yes")
    wb_    = sum(1 for p in papers if (p.get("wb_affiliation_auto") or "") == "Yes")
    mult   = sum(1 for p in papers if p.get("multilateral_affiliation"))
    af_any = sum(1 for p in papers if p.get("is_any_author_africa"))
    af_1st = sum(1 for p in papers if p.get("is_first_author_africa"))
    af_str = sum(1 for p in papers if p.get("is_africa_institution_strict"))
    unk    = sum(1 for p in papers if p.get("geography_clean") == "Unclassified")
    r3     = sum(1 for p in papers if (p.get("relevance_score") or 0) == 3)
    r2     = sum(1 for p in papers if (p.get("relevance_score") or 0) == 2)
    tA     = sum(1 for p in papers if p.get("match_tier") == "A")
    tB     = sum(1 for p in papers if p.get("match_tier") == "B")
    tC     = sum(1 for p in papers if p.get("match_tier") == "C")
    tAND   = sum(1 for p in papers if p.get("match_tier") == "AND")

    fy0        = completed_fys[0]
    fy0_total  = sum(1 for p in papers if p.get("fy")==fy0)
    af_any_fy0 = sum(1 for p in papers if p.get("fy")==fy0 and p.get("is_any_author_africa"))
    af_1st_fy0 = sum(1 for p in papers if p.get("fy")==fy0 and p.get("is_first_author_africa"))

    pct     = lambda n: (n / total) if total else 0
    pct_fy0 = lambda n: (n / fy0_total) if fy0_total else 0

    # ── Column headers row 1 ──────────────────────────────────────────────────
    for c, h in enumerate(["Metric", "Note", "Count", "Share"], 1):
        _cell(ws, 1, c, h, bold=True, white_text=True, fill=_FILL_DARK)
    ws.cell(1,3).alignment = _AL_CENTER
    ws.cell(1,4).alignment = _AL_CENTER
    r = 2

    # ── TOTALS ────────────────────────────────────────────────────────────────
    _section_row(ws, r, "── TOTALS (analysed set only: relevance score ≥ 2) ──"); r+=1
    _data_row(ws, r, "TOTAL PAPERS ANALYSED", "all figures below refer to this set",
              total, bold_label=True); r+=1
    _data_row(ws, r, "  Excluded — score 1 (no LSMS signal in title/abstract)",
              "in 'Not Relevant (Backup)' sheet", n1); r+=1
    _data_row(ws, r, "  Excluded — score 0 (review / meta-analysis)",
              "in 'Not Relevant (Backup)' sheet", n0); r+=1
    _data_row(ws, r, "  Total retrieved before exclusion", "", total+n0+n1); r+=1
    _data_row(ws, r, "Peer-reviewed journal articles (auto-detected)", "", peer, pct(peer)); r+=1
    _data_row(ws, r, "World Bank–affiliated papers (auto-detected)", "", wb_, pct(wb_)); r+=1
    _data_row(ws, r, "Multilateral org–affiliated (IFPRI/FAO/CGIAR…)", "", mult, pct(mult)); r+=1
    r += 1   # blank

    # ── FLOW ─────────────────────────────────────────────────────────────────
    _section_row(ws, r, "── FLOW: new papers by WBG fiscal year ──────────────"); r+=1
    flow_cur = sum(1 for p in papers if p.get("fy") == current_fy)
    _data_row(ws, r, f"  {current_fy}", "current FY — still in progress", flow_cur); r+=1
    for i, fy in enumerate(completed_fys):
        n = sum(1 for p in papers if p.get("fy") == fy)
        note = "◄ most recently completed FY" if i == 0 else ""
        # Highlight FY26 (most recently completed)
        _data_row(ws, r, f"  {fy}", note, n, hl_count=(i==0)); r+=1
    r += 1   # blank

    # ── MATCH TIERS ──────────────────────────────────────────────────────────
    _section_row(ws, r, "── HOW EACH PAPER WAS MATCHED ─────────────────────"); r+=1
    _data_row(ws, r, "  Tier A — unambiguous survey name (e.g. 'Uganda National Panel Survey')",
              "full-text match accepted", tA, pct(tA)); r+=1
    _data_row(ws, r, "  Tier B — medium identifier (e.g. 'LSMS-ISA', 'Tanzania NPS')",
              "required in title/abstract", tB, pct(tB)); r+=1
    _data_row(ws, r, "  Tier C — short acronym (e.g. 'IHPS', 'ESS1')",
              "case-sensitive + country context + discipline", tC, pct(tC)); r+=1
    _data_row(ws, r, "  Compound AND (e.g. 'HFPS and Burkina Faso')",
              "both components required", tAND, pct(tAND)); r+=1
    r += 1

    # ── RELEVANCE ─────────────────────────────────────────────────────────────
    _section_row(ws, r, "── RELEVANCE SCORE — composition of the analysed set ──"); r+=1
    _data_row(ws, r, "  Score 3 — survey name appears in the paper title",
              "strongest data-use signal", r3, pct(r3)); r+=1
    _data_row(ws, r, "  Score 2 — survey in abstract, data-use language, or LSMS named",
              "", r2, pct(r2)); r+=1
    _data_row(ws, r, "  (Scores 0 and 1 are EXCLUDED — see 'Not Relevant (Backup)' sheet)",
              "not counted anywhere above"); r+=1
    r += 1

    # ── SHARE ─────────────────────────────────────────────────────────────────
    _section_row(ws, r, "── SHARE: Geography of Authors ──────────────────────────"); r+=1
    _data_row(ws, r, "NOTE: institution-based (undercounts African diaspora at WB/US/EU)"); r+=1
    _data_row(ws, r, "  Any author at African institution", "≈ original SSA + Mixed",
              af_any, pct(af_any), hl_share=True); r+=1
    _data_row(ws, r, f"    of which in {fy0} (share of {fy0} papers)", f"{af_any_fy0} of {fy0_total} {fy0} papers",
              af_any_fy0, pct_fy0(af_any_fy0), hl_share=True); r+=1
    _data_row(ws, r, "  First author at African institution", "", af_1st, pct(af_1st)); r+=1
    _data_row(ws, r, f"    of which in {fy0} (share of {fy0} papers)", f"{af_1st_fy0} of {fy0_total} {fy0} papers",
              af_1st_fy0, pct_fy0(af_1st_fy0), hl_share=True); r+=1
    _data_row(ws, r, "  ALL authors at SSA institution", "", af_str, pct(af_str)); r+=1
    _data_row(ws, r, "  Geography unclassified (no OpenAlex institution data)", "",
              unk, pct(unk)); r+=1

    # Format all share column cells
    for row_i in range(2, r+1):
        cell = ws.cell(row_i, 4)
        if isinstance(cell.value, float):
            cell.number_format = "0%"

    # ── Column widths (matching reference file) ────────────────────────────────
    ws.column_dimensions["A"].width = 75
    ws.column_dimensions["B"].width = 48
    ws.column_dimensions["C"].width = 23
    ws.column_dimensions["D"].width = 23

    # ── FY Trend table (below metrics, same sheet) ────────────────────────────
    tbl_start = r + 2
    for j, h in enumerate(["Fiscal Year","Papers",
                            "Africa share (any author)","Africa share (1st author)"], 1):
        c = ws.cell(tbl_start, j, h)
        c.font = _F_WHITE_BOLD
        c.fill = _FILL_DARK
        c.alignment = _AL_CENTER

    chart_rows = []
    for _, row_d in trend_df.iterrows():
        fy    = str(row_d["FY"])
        n_all = int(row_d["All Papers"])
        n_aa  = int(row_d["Africa (any author)"])
        n_fa  = int(row_d["Africa (1st author)"])
        if n_all == 0 or fy == current_fy:
            continue
        r_i = tbl_start + 1 + len(chart_rows)
        ws.cell(r_i, 1, fy)
        ws.cell(r_i, 2, n_all)
        ws.cell(r_i, 3, n_aa / n_all).number_format = "0%"
        ws.cell(r_i, 4, n_fa / n_all).number_format = "0%"
        chart_rows.append((fy, n_all, n_aa, n_fa))
    tbl_end = tbl_start + len(chart_rows)

    if not chart_rows:
        return

    # ── CHART 1: Bar — papers per FY ─────────────────────────────────────────
    cats  = Reference(ws, min_col=1, min_row=tbl_start+1, max_row=tbl_end)
    tot_r = Reference(ws, min_col=2, max_col=2, min_row=tbl_start, max_row=tbl_end)

    bar = BarChart()
    bar.type = "col"
    bar.grouping = "clustered"
    bar.title = "LSMS Papers per Fiscal Year (number of papers)"
    bar.y_axis.title = "Number of papers"
    bar.x_axis.title = "Fiscal Year (World Bank: 1 Jul – 30 Jun)"
    bar.x_axis.tickLblPos = "low"
    bar.y_axis.crosses = "autoZero"
    bar.width, bar.height = 15, 7.5   # exact match to reference
    bar.legend = None
    bar.add_data(tot_r, titles_from_data=True)
    bar.set_categories(cats)
    bar.series[0].graphicalProperties.solidFill = C_DARK
    bar.series[0].graphicalProperties.line.solidFill = C_DARK
    bar.dataLabels = DataLabelList()
    bar.dataLabels.showVal = True
    bar.dataLabels.showLegendKey = False
    bar.dataLabels.showCatName  = False
    bar.dataLabels.showSerName  = False
    # Anchor at E1 (col=4 offset, row=0 offset) — matches reference file
    ws.add_chart(bar, "E1")

    # ── CHART 2: Line with markers — Africa share per FY ─────────────────────
    any_r = Reference(ws, min_col=3, max_col=3, min_row=tbl_start, max_row=tbl_end)
    fst_r = Reference(ws, min_col=4, max_col=4, min_row=tbl_start, max_row=tbl_end)

    line = LineChart()
    line.title = "Share of African Scholars per Fiscal Year"
    line.y_axis.title = "Share of papers (%)"
    line.x_axis.title = "Fiscal Year (World Bank: 1 Jul – 30 Jun)"
    line.x_axis.tickLblPos = "low"
    line.y_axis.numFmt = "0%"
    line.y_axis.crosses = "autoZero"
    line.width, line.height = 15, 7.5   # exact match to reference
    line.style = 10
    line.add_data(any_r, titles_from_data=True)
    line.add_data(fst_r, titles_from_data=True)
    line.set_categories(cats)
    COLORS = [("2E7D32", "circle"), ("E65100", "diamond")]
    for idx, (clr, sym) in enumerate(COLORS):
        s = line.series[idx]
        s.graphicalProperties.line.solidFill = clr
        s.graphicalProperties.line.width = 28000
        s.marker.symbol, s.marker.size = sym, 6
        s.marker.graphicalProperties.solidFill = clr
        s.marker.graphicalProperties.line.solidFill = clr
    # Each point shows only its percentage value — showing the series name or
    # category name here as well is what causes the labels to overlap into an
    # unreadable block, since every point would carry the full series name.
    line.dataLabels = DataLabelList()
    line.dataLabels.showVal = True
    line.dataLabels.numFmt = "0%"
    line.dataLabels.showLegendKey = False
    line.dataLabels.showCatName  = False
    line.dataLabels.showSerName  = False
    # Compact legend inside the plot area, top-left, so the two colours/markers
    # are identified once instead of repeated on every data point.
    line.legend = Legend()
    line.legend.position = "tr"
    line.legend.overlay = True
    line.legend.layout = Layout(
        manualLayout=ManualLayout(
            xMode="edge", yMode="edge",
            x=0.02, y=0.02, h=0.16, w=0.40,
        )
    )
    # Anchor at E24 (row=23 offset) — matches reference file
    ws.add_chart(line, "E24")

    # ── CHART 3: Pie — journal tier distribution ─────────────────────────────
    _TIER_ORDER = [
        "1 — Top General Econ",
        "2 — Top Field",
        "3 — Quality Field",
        "4 — Other Peer-Reviewed",
        "WP — Working Paper / Non-Journal",
    ]
    tier_counts = {t: 0 for t in _TIER_ORDER}
    other = 0
    for p in papers:
        jt = p.get("journal_tier") or ""
        if jt in tier_counts:
            tier_counts[jt] += 1
        elif jt:
            other += 1
    # Data table for pie, placed a few rows below the FY trend table
    pie_start = tbl_end + 3
    ws.cell(pie_start, 1, "Journal Tier").font = _F_WHITE_BOLD
    ws.cell(pie_start, 1).fill = _FILL_DARK
    ws.cell(pie_start, 2, "Papers").font = _F_WHITE_BOLD
    ws.cell(pie_start, 2).fill = _FILL_DARK
    pie_labels = [
        "Tier 1 — Top General Econ",
        "Tier 2 — Top Field",
        "Tier 3 — Quality Field",
        "Tier 4 — Other Peer-Reviewed",
        "Working Paper / Non-Journal",
    ]
    for k, (tkey, lbl) in enumerate(zip(_TIER_ORDER, pie_labels), 1):
        ws.cell(pie_start + k, 1, lbl)
        ws.cell(pie_start + k, 2, tier_counts[tkey])
    pie_end = pie_start + len(_TIER_ORDER)

    pie = PieChart()
    pie.title = "Papers by Journal Tiers"
    pie.width, pie.height = 15, 7.5
    pie_data = Reference(ws, min_col=2, min_row=pie_start, max_row=pie_end)
    pie_cats = Reference(ws, min_col=1, min_row=pie_start + 1, max_row=pie_end)
    pie.add_data(pie_data, titles_from_data=True)
    pie.set_categories(pie_cats)
    # Raw paper counts on each slice (not percentages)
    pie.dataLabels = DataLabelList()
    pie.dataLabels.showPercent = False
    pie.dataLabels.showVal = True
    # Monochrome navy-to-pale-blue gradient, darkest = highest tier, instead
    # of openpyxl's default rainbow palette.
    _PIE_BLUES = ["1F3864", "2E75B6", "5B9BD5", "9DC3E6", "DEEBF7"]
    for idx, clr in enumerate(_PIE_BLUES):
        pt = DataPoint(idx=idx)
        pt.graphicalProperties.solidFill = clr
        pie.series[0].data_points.append(pt)
    # Anchor below the two FY charts
    ws.add_chart(pie, "E47")


# ─────────────────────────────────────────────────────────────────────────────
# CELL SANITIZER  (openpyxl rejects XML control chars and HTML entities)
# ─────────────────────────────────────────────────────────────────────────────

_ILLEGAL_XML = re.compile(r'[\x00-\x08\x0b\x0c\x0e-\x1f]')

def _clean_cell(v):
    """Decode HTML entities, strip XML control chars, cap at Excel's 32,767-char limit."""
    if not isinstance(v, str):
        return v
    v = html_module.unescape(v)          # &lt; → <   &amp; → &   etc.
    v = _ILLEGAL_XML.sub('', v)          # remove control chars Excel forbids
    return v[:32000] + '…' if len(v) > 32000 else v


def export_excel(papers: list, review: list, search_log: list, output_path: str,
                 min_year: int = 1980, excluded: Optional[list] = None):
    """
    Sheet order:
      1. Papers                  - the analysed set
      2. Analysis                - metrics + three charts
      3. FY Trend                - data backing the charts
      4. Keywords                - all search terms, tiers, queries
      5. Search Log              - audit trail
      6. Dedup Review            - fuzzy title matches (if any)
      7. Not Relevant (Backup)   - LAST. No LSMS signal in title/abstract.
    """
    if not papers:
        print("[export] No papers to write.")
        return

    def _fy_key(p):
        # Newest FY first. FY27 -> 2027, FY99 -> 1999, so FY27 sorts above FY99.
        return (-_fy_to_year(p.get("fy") or ""), (p.get("title") or "").lower())

    df = pd.DataFrame(sorted(papers, key=_fy_key))
    for c in OUTPUT_COLS:
        if c not in df.columns:
            df[c] = None
    df = df[[c for c in OUTPUT_COLS if c in df.columns] +
            [c for c in df.columns if c not in OUTPUT_COLS and not c.startswith("_")]]

    def _safe(v):
        if not isinstance(v, str): return v
        v = html_module.unescape(v)
        v = _ILLEGAL_XML.sub('', v)
        if v and v[0] in ('=', '+', '-', '@'): v = ' ' + v
        return v[:32000] + '\u2026' if len(v) > 32000 else v

    try:    df = df.map(_safe)
    except: df = df.applymap(_safe)

    current_fy, completed_fys = current_and_prior_fy()
    trend_df = _build_trend_df(papers, current_fy)

    # Keywords sheet
    from keywords import SURVEY_FAMILIES as _KW
    _TIER_DESC = {
        "A":   "Unambiguous - full-text match accepted",
        "B":   "Medium - required in title or abstract",
        "C":   "Short acronym - case-sensitive + country context + discipline",
        "AND": "Boolean AND - both components required",
    }
    kw_rows = []
    for f in _KW:
        hints = ", ".join(f.get("context_hints") or [])
        for term, tier in f["terms"]:
            kw_rows.append((f["label"], f["region"], term, tier, _TIER_DESC[tier],
                            _build_search_query(term, tier), hints if tier == "C" else ""))
    kw_df = pd.DataFrame(kw_rows, columns=[
        "Survey Family", "Region", "Search Term", "Match Tier",
        "Matching Rule", "OpenAlex Query Sent", "Required Context Words"])

    PAPERS, ANALYSIS, TREND = "Papers", "Analysis", "FY Trend"
    BACKUP = "Not Relevant (Backup)"
    hdr_fill = PatternFill("solid", start_color="1F5C99", end_color="1F5C99")
    hdr_font = Font(bold=True, color="FFFFFF")

    def _style(ws):
        for cell in ws[1]:
            cell.font = hdr_font
            cell.fill = hdr_fill
            cell.alignment = Alignment(horizontal="center", wrap_text=True)

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        # Papers first
        df.to_excel(writer, sheet_name=PAPERS, index=False)
        # Analysis: create as empty sheet, fill in below after writer closes
        pd.DataFrame().to_excel(writer, sheet_name=ANALYSIS)
        trend_df.to_excel(writer, sheet_name=TREND, index=False)
        kw_df.to_excel(writer, sheet_name="Keywords", index=False)
        if search_log:
            pd.DataFrame(search_log).to_excel(writer, sheet_name="Search Log", index=False)
        if review:
            pd.DataFrame(review).to_excel(writer, sheet_name="Dedup Review", index=False)
        # BACKUP LAST
        if excluded:
            ex_df = pd.DataFrame(sorted(excluded, key=_fy_key))
            for c in OUTPUT_COLS:
                if c not in ex_df.columns: ex_df[c] = None
            ex_df = ex_df[[c for c in OUTPUT_COLS if c in ex_df.columns]]
            try:    ex_df = ex_df.map(_safe)
            except: ex_df = ex_df.applymap(_safe)
            ex_df.to_excel(writer, sheet_name=BACKUP, index=False)

        # Papers sheet styling
        ws_p = writer.sheets[PAPERS]
        ws_p.freeze_panes = "A2"
        _style(ws_p)
        for i, cn in enumerate(df.columns, 1):
            ws_p.column_dimensions[get_column_letter(i)].width = COL_WIDTHS.get(cn, 13)

        # Backup styling
        if excluded and BACKUP in writer.sheets:
            ws_b = writer.sheets[BACKUP]
            ws_b.freeze_panes = "A2"
            _style(ws_b)
            for i, cn in enumerate(ex_df.columns, 1):
                ws_b.column_dimensions[get_column_letter(i)].width = COL_WIDTHS.get(cn, 13)

        # Keywords
        ws_k = writer.sheets["Keywords"]
        ws_k.freeze_panes = "A2"
        _style(ws_k)
        for ltr, w in {"A":34,"B":20,"C":58,"D":11,"E":52,"F":42,"G":30}.items():
            ws_k.column_dimensions[ltr].width = w

        # FY Trend
        ws_t = writer.sheets[TREND]
        _style(ws_t)
        for ltr, w in {"A":8,"B":14,"C":20,"D":20,"E":16,"F":16}.items():
            ws_t.column_dimensions[ltr].width = w

    # ── Analysis sheet: re-open and write with exact styling ──────────────────
    # (done after the writer closes so we have full control without pandas
    #  interfering with cell styles)
    from openpyxl import load_workbook
    wb_out = load_workbook(output_path)
    ws_a = wb_out[ANALYSIS]
    # Delete the empty placeholder row pandas wrote
    ws_a.delete_rows(1, ws_a.max_row)
    write_analysis_sheet(ws_a, papers, current_fy, completed_fys,
                         min_year, excluded, trend_df)
    wb_out.save(output_path)

    print(f"[export] {len(df)} papers -> {output_path}")
    if excluded:
        print(f"[export] {len(excluded)} no-signal papers -> '{BACKUP}' sheet")
    if review:
        print(f"[export] {len(review)} fuzzy -> 'Dedup Review'")


# ─────────────────────────────────────────────────────────────────────────────
# LOAD EXISTING
# ─────────────────────────────────────────────────────────────────────────────

def load_existing(path: str) -> pd.DataFrame:
    p = Path(path)
    if not p.exists():
        print(f"[warn] --merge-existing file not found: {path}")
        return pd.DataFrame()
    if p.suffix.lower() == ".csv":
        return pd.read_csv(path)
    for sheet in ("Papers", "List of pubs."):
        try:
            return pd.read_excel(path, sheet_name=sheet)
        except Exception:
            continue
    return pd.DataFrame()

# ─────────────────────────────────────────────────────────────────────────────
# MAIN
# ─────────────────────────────────────────────────────────────────────────────

def run_discovery(
    api_key: str = "",
    since_fy: Optional[str] = None,
    min_year: int = 1980,
    min_relevance_score: int = 0,   # 0 = include all keyword matches (recommended)
    use_crossref: bool = False,
    merge_existing: Optional[str] = None,
    output_path: Optional[str] = None,
    fuzzy_threshold: float = 0.88,
    verbose: bool = True,
    test_mode: bool = False,
) -> list:

    since_date = None
    if not api_key:
        print("[warn] No --api-key supplied. Free budget is only 100 searches/day without a key.")
        print("[warn] Get a FREE key in 30s at: https://openalex.org/settings/api", flush=True)
    if since_fy:
        since_date = fy_start_date(since_fy).isoformat()
        if verbose:
            print(f"[info] Filtering papers ≥ {since_date} ({since_fy} start)")

    if output_path is None:
        sfx = f"_since_{since_fy}" if since_fy else ""
        output_path = f"LSMS_papers{sfx}_{__import__('datetime').datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

    existing_df = load_existing(merge_existing) if merge_existing else pd.DataFrame()

    oa_fetcher = OpenAlexFetcher(api_key=api_key)
    cr_fetcher = CrossrefFetcher(api_key=api_key) if use_crossref else None

    all_papers: list = []
    search_log: list = []

    families = SURVEY_FAMILIES[:2] if test_mode else SURVEY_FAMILIES

    n_families = len(families)
    for f_idx, family in enumerate(families, 1):
        label = family["label"]

        # Search each term individually — prints its own live progress
        oa_results = oa_fetcher.search_family(
            family, since_date,
            family_index=f_idx,
            total_families=n_families,
            verbose=verbose,
        )
        all_papers.extend(oa_results)
        search_log.append({
            "survey_family": label,
            "source": "OpenAlex",
            "n_terms": len(family["terms"]),
            "n_results_raw": len(oa_results),
            "run_date": date.today().isoformat(),
            "since_date": since_date or "all",
        })

        if verbose:
            running_total = len(all_papers)
            print(f"  Running total (raw): {running_total:,}", flush=True)

        # Optional Crossref supplement (long terms, title-only)
        if cr_fetcher:
            cr_total = 0
            for term, _tier in family["terms"]:
                cr = cr_fetcher.search_term(term, family)
                all_papers.extend(cr)
                cr_total += len(cr)
            if cr_total and verbose:
                print(f"  CR supplement: +{cr_total}")
            search_log.append({
                "survey_family": label,
                "source": "Crossref",
                "n_terms": len(family["terms"]),
                "n_results_raw": cr_total,
                "run_date": date.today().isoformat(),
                "since_date": since_date or "all",
            })

    # Drop pre-1980 papers and impossible future publication years (OpenAlex
    # occasionally carries a placeholder date ahead of the present). One
    # extra year of headroom covers legitimately forthcoming articles.
    CURRENT_YEAR = date.today().year
    MAX_VALID_YEAR = CURRENT_YEAR + 1
    n_before = len(all_papers)
    all_papers = [p for p in all_papers
                 if min_year <= (p.get("year") or 0) <= MAX_VALID_YEAR]
    if verbose and len(all_papers) < n_before:
        print(f"  Dropped {n_before - len(all_papers)} papers outside the valid "
              f"year range ({min_year}\u2013{MAX_VALID_YEAR})", flush=True)

    # Dedup runs before the relevance gate: a paper found via several
    # independent keyword matches needs to be merged (and multi-family
    # boosted) before scoring, or the duplicates get scored — and possibly
    # excluded — in isolation.
    if verbose:
        print(f"\n[dedup] {len(all_papers)} raw → deduplicating...")

    clean_all, review = deduplicate(all_papers, existing_df, fuzzy_threshold)

    # Only papers scoring >= 2 enter the analysed set (survey named in the
    # abstract/title, explicit data-use language, LSMS/World Bank named,
    # empirical vocabulary, a Tier-A match with no abstract to check, or a
    # multi-family match). Scores 0 (review/meta-analysis) and 1 (fulltext-only
    # match, no real signal) go to the 'Not Relevant (Backup)' sheet instead.
    RELEVANCE_CUTOFF = max(2, min_relevance_score)
    excluded_low_relevance: list = []
    kept: list = []
    for p in clean_all:
        if (p.get("relevance_score") or 0) >= RELEVANCE_CUTOFF:
            kept.append(p)
        else:
            excluded_low_relevance.append(p)
    clean = kept
    if verbose and excluded_low_relevance:
        n0 = sum(1 for p in excluded_low_relevance if (p.get("relevance_score") or 0) == 0)
        n1 = sum(1 for p in excluded_low_relevance if (p.get("relevance_score") or 0) == 1)
        print(f"  Set aside {len(excluded_low_relevance):,} papers "
              f"(score 1: {n1:,} no LSMS signal | score 0: {n0:,} reviews) "
              f"-> 'Not Relevant (Backup)' sheet", flush=True)

    if verbose:
        current_fy, completed_fys = current_and_prior_fy()
        total = len(clean)
        no_fy = sum(1 for p in clean if not p.get("fy"))
        af_fa  = sum(1 for p in clean if p.get("is_first_author_africa"))
        af_any = sum(1 for p in clean if p.get("is_any_author_africa"))
        af_str = sum(1 for p in clean if p.get("is_africa_institution_strict"))
        r3     = sum(1 for p in clean if (p.get("relevance_score") or 0) >= 3)
        r2     = sum(1 for p in clean if (p.get("relevance_score") or 0) >= 2)
        print(f"\n=== RESULTS ===")
        print(f"Unique papers (≥ 1980):            {total:,}")
        for fy in completed_fys:
            n = sum(1 for p in clean if p.get("fy") == fy)
            tag = "  ← most recently completed" if fy == completed_fys[0] else ""
            print(f"FLOW {fy}: {n}{tag}")
        print(f"FLOW {current_fy} (in progress): {sum(1 for p in clean if p.get('fy') == current_fy)}")
        if total:
            print(f"SHARE Africa (first author):        {af_fa}/{total} = {af_fa/total:.1%}")
            print(f"SHARE Africa (any author):          {af_any}/{total} = {af_any/total:.1%}")
            print(f"SHARE Africa (strict/all-SSA):      {af_str}/{total} = {af_str/total:.1%}")
            print(f"Relevance score ≥3 (survey in title): {r3:,} = {r3/total:.1%}")
            print(f"Relevance score ≥2 (likely data use): {r2:,} = {r2/total:.1%}")
        print(f"Papers without month (FY blank):    {no_fy}")
        print(f"Fuzzy review candidates:            {len(review)}")

    export_excel(clean, review, search_log, output_path, min_year=min_year,
                 excluded=excluded_low_relevance)
    return clean


def main():
    ap = argparse.ArgumentParser(
        description=__doc__,
        formatter_class=argparse.RawDescriptionHelpFormatter,
    )
    ap.add_argument("--api-key", default="",
                    help="OpenAlex API key (FREE at openalex.org/settings/api). Required for >100 searches/day.")
    ap.add_argument("--since-fy", metavar="FYxx",
                    help="Only fetch papers published since this FY start, e.g. FY24")
    ap.add_argument("--min-relevance", type=int, default=0,
                    help="Minimum relevance score (0=include all [default], 1=exclude likely reviews, 2=require explicit data-use language)")
    ap.add_argument("--min-year", type=int, default=1980,
                    help="Exclude papers before this year (default: 1980, when LSMS was established)")
    ap.add_argument("--merge-existing", metavar="FILE",
                    help="Path to existing Excel/CSV; new papers deduped against it")
    ap.add_argument("--output", metavar="FILE",
                    help="Output Excel path (default: LSMS_papers_YYYYMMDD.xlsx)")
    ap.add_argument("--crossref", action="store_true",
                    help="Also search Crossref (title-only, max 100/term). Slower.")
    ap.add_argument("--fuzzy-threshold", type=float, default=0.88)
    ap.add_argument("--test", action="store_true",
                    help="Quick test: 2 families only")
    ap.add_argument("-q", "--quiet", action="store_true")
    args = ap.parse_args()

    run_discovery(
        api_key=args.api_key,
        min_year=args.min_year,
        min_relevance_score=args.min_relevance,
        since_fy=args.since_fy,
        use_crossref=args.crossref,
        merge_existing=args.merge_existing,
        output_path=args.output,
        fuzzy_threshold=args.fuzzy_threshold,
        verbose=not args.quiet,
        test_mode=args.test,
    )


if __name__ == "__main__":
    main()
